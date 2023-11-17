# 23/05作成
# 通常歩行を100%として正規化
# 筋肉ごとに1歩目から4歩目までの合計を算出

import pandas as pd
import numpy as np
import glob
import os
import statistics
import openpyxl
from tqdm import trange, tqdm

n = 4 # 試行数（ファイルの数）
condition = "condition4"

#空リスト
L_Bic_1 = []
R_Bic_1 = []
L_Tri_1 = []
R_Tri_1 = []
L_Vlo_1 = []
R_Vlo_1 = []
L_Rec_1 = []
R_Rec_1 = []
L_Tib_1 = []
R_Tib_1 = []
L_Gas_1 = []
R_Gas_1 = []

print("第3_1相処理中...")
files = glob.glob("phase3_1/" + condition + "/*.csv")

for file in tqdm(files):
    df = pd.read_csv(file)
    df = df.abs()

    list = df["L.Biceps Br.(uV):"]
    L_Bic_1.extend(list)
    list = df["R.Biceps Br.(uV):"]
    R_Bic_1.extend(list)
    list = df["L.Med. Triceps(uV):"]
    L_Tri_1.extend(list)
    list = df["R.Med. Triceps(uV):"]
    R_Tri_1.extend(list)
    list = df["L.Vlo(uV):"]
    L_Vlo_1.extend(list)
    list = df["R.Vlo(uV):"]
    R_Vlo_1.extend(list)
    list = df["L.Rectus Fem.(uV):"]
    L_Rec_1.extend(list)
    list = df["R.Rectus Fem.(uV):"]
    R_Rec_1.extend(list)
    list = df["L.Tib.Ant.(uV):"]
    L_Tib_1.extend(list)
    list = df["R.Tib.Ant.(uV):"]
    R_Tib_1.extend(list)
    list = df["L.Med. Gastro(uV):"]
    L_Gas_1.extend(list)
    list = df["R.Med. Gastro(uV):"]
    R_Gas_1.extend(list)

#空リスト
L_Bic_2 = []
R_Bic_2 = []
L_Tri_2 = []
R_Tri_2 = []
L_Vlo_2 = []
R_Vlo_2 = []
L_Rec_2 = []
R_Rec_2 = []
L_Tib_2 = []
R_Tib_2 = []
L_Gas_2 = []
R_Gas_2 = []

print("第3_2相処理中...")
files = glob.glob("phase3_2/" + condition + "/*.csv")

for file in tqdm(files):
    df = pd.read_csv(file)
    df = df.abs()

    list = df["L.Biceps Br.(uV):"]
    L_Bic_2.extend(list)
    list = df["R.Biceps Br.(uV):"]
    R_Bic_2.extend(list)
    list = df["L.Med. Triceps(uV):"]
    L_Tri_2.extend(list)
    list = df["R.Med. Triceps(uV):"]
    R_Tri_2.extend(list)
    list = df["L.Vlo(uV):"]
    L_Vlo_2.extend(list)
    list = df["R.Vlo(uV):"]
    R_Vlo_2.extend(list)
    list = df["L.Rectus Fem.(uV):"]
    L_Rec_2.extend(list)
    list = df["R.Rectus Fem.(uV):"]
    R_Rec_2.extend(list)
    list = df["L.Tib.Ant.(uV):"]
    L_Tib_2.extend(list)
    list = df["R.Tib.Ant.(uV):"]
    R_Tib_2.extend(list)
    list = df["L.Med. Gastro(uV):"]
    L_Gas_2.extend(list)
    list = df["R.Med. Gastro(uV):"]
    R_Gas_2.extend(list)

#空リスト
L_Bic_3 = []
R_Bic_3 = []
L_Tri_3 = []
R_Tri_3 = []
L_Vlo_3 = []
R_Vlo_3 = []
L_Rec_3 = []
R_Rec_3 = []
L_Tib_3 = []
R_Tib_3 = []
L_Gas_3 = []
R_Gas_3 = []

print("第3_3相処理中...")
files = glob.glob("phase3_3/" + condition + "/*.csv")

for file in tqdm(files):
    df = pd.read_csv(file)
    df = df.abs()

    list = df["L.Biceps Br.(uV):"]
    L_Bic_3.extend(list)
    list = df["R.Biceps Br.(uV):"]
    R_Bic_3.extend(list)
    list = df["L.Med. Triceps(uV):"]
    L_Tri_3.extend(list)
    list = df["R.Med. Triceps(uV):"]
    R_Tri_3.extend(list)
    list = df["L.Vlo(uV):"]
    L_Vlo_3.extend(list)
    list = df["R.Vlo(uV):"]
    R_Vlo_3.extend(list)
    list = df["L.Rectus Fem.(uV):"]
    L_Rec_3.extend(list)
    list = df["R.Rectus Fem.(uV):"]
    R_Rec_3.extend(list)
    list = df["L.Tib.Ant.(uV):"]
    L_Tib_3.extend(list)
    list = df["R.Tib.Ant.(uV):"]
    R_Tib_3.extend(list)
    list = df["L.Med. Gastro(uV):"]
    L_Gas_3.extend(list)
    list = df["R.Med. Gastro(uV):"]
    R_Gas_3.extend(list)

#空リスト
L_Bic_4 = []
R_Bic_4 = []
L_Tri_4 = []
R_Tri_4 = []
L_Vlo_4 = []
R_Vlo_4 = []
L_Rec_4 = []
R_Rec_4 = []
L_Tib_4 = []
R_Tib_4 = []
L_Gas_4 = []
R_Gas_4 = []

print("第3_4相処理中...")
files = glob.glob("phase3_4/" + condition + "/*.csv")

for file in tqdm(files):
    df = pd.read_csv(file)
    df = df.abs()

    list = df["L.Biceps Br.(uV):"]
    L_Bic_4.extend(list)
    list = df["R.Biceps Br.(uV):"]
    R_Bic_4.extend(list)
    list = df["L.Med. Triceps(uV):"]
    L_Tri_4.extend(list)
    list = df["R.Med. Triceps(uV):"]
    R_Tri_4.extend(list)
    list = df["L.Vlo(uV):"]
    L_Vlo_4.extend(list)
    list = df["R.Vlo(uV):"]
    R_Vlo_4.extend(list)
    list = df["L.Rectus Fem.(uV):"]
    L_Rec_4.extend(list)
    list = df["R.Rectus Fem.(uV):"]
    R_Rec_4.extend(list)
    list = df["L.Tib.Ant.(uV):"]
    L_Tib_4.extend(list)
    list = df["R.Tib.Ant.(uV):"]
    R_Tib_4.extend(list)
    list = df["L.Med. Gastro(uV):"]
    L_Gas_4.extend(list)
    list = df["R.Med. Gastro(uV):"]
    R_Gas_4.extend(list)

# 全リスト結合
L_Bic_all = L_Bic_1 + L_Bic_2 + L_Bic_3 + L_Bic_4
R_Bic_all = R_Bic_1 + R_Bic_2 + R_Bic_3 + R_Bic_4
L_Tri_all = L_Tri_1 + L_Tri_2 + L_Tri_3 + L_Tri_4
R_Tri_all = R_Tri_1 + R_Tri_2 + R_Tri_3 + R_Tri_4
L_Vlo_all = L_Vlo_1 + L_Vlo_2 + L_Vlo_3 + L_Vlo_4
R_Vlo_all = R_Vlo_1 + R_Vlo_2 + R_Vlo_3 + R_Vlo_4
L_Rec_all = L_Rec_1 + L_Rec_2 + L_Rec_3 + L_Rec_4
R_Rec_all = R_Rec_1 + R_Rec_2 + R_Rec_3 + R_Rec_4
L_Tib_all = L_Tib_1 + L_Tib_2 + L_Tib_3 + L_Tib_4
R_Tib_all = R_Tib_1 + R_Tib_2 + R_Tib_3 + R_Tib_4
L_Gas_all = L_Gas_1 + L_Gas_2 + L_Gas_3 + L_Gas_4
R_Gas_all = R_Gas_1 + R_Gas_2 + R_Gas_3 + R_Gas_4

# エクセル作成
wbname = "EMG_ave_" + condition + ".xlsx"
wb = openpyxl.Workbook()
wb.save("excel/" + wbname)

book = openpyxl.load_workbook("excel/" + wbname)
sheet = book["Sheet"]

for j in range(0, 12):
    excel_list = ["左上腕二頭筋", "右上腕二頭筋", "左上腕三頭筋", "右上腕三頭筋", "左大腿直筋", "右大腿直筋",
                    "左外側広筋", "右外側広筋", "左前脛骨筋", "右前脛骨筋", "左腓腹筋内側頭", "右腓腹筋内側頭"]
    sheet.cell(1, 1 + j).value = excel_list[j]

# セルに書込
# 上腕二頭筋
sheet.cell(2, 1).value = statistics.mean(L_Bic_all)
sheet.cell(3, 1).value = statistics.stdev(L_Bic_all)
sheet.cell(2, 2).value = statistics.mean(R_Bic_all)
sheet.cell(3, 2).value = statistics.stdev(R_Bic_all)
# 上腕三頭筋
sheet.cell(2, 3).value = statistics.mean(L_Tri_all)
sheet.cell(3, 3).value = statistics.stdev(L_Tri_all)
sheet.cell(2, 4).value = statistics.mean(R_Tri_all)
sheet.cell(3, 4).value = statistics.stdev(R_Tri_all)
# 大腿直筋
sheet.cell(2, 5).value = statistics.mean(L_Rec_all)
sheet.cell(3, 5).value = statistics.stdev(L_Rec_all)
sheet.cell(2, 6).value = statistics.mean(R_Rec_all)
sheet.cell(3, 6).value = statistics.stdev(R_Rec_all)
# 外側広筋
sheet.cell(2, 7).value = statistics.mean(L_Vlo_all)
sheet.cell(3, 7).value = statistics.stdev(L_Vlo_all)
sheet.cell(2, 8).value = statistics.mean(R_Vlo_all)
sheet.cell(3, 8).value = statistics.stdev(R_Vlo_all)
# 前脛骨筋
sheet.cell(2, 9).value = statistics.mean(L_Tib_all)
sheet.cell(3, 9).value = statistics.stdev(L_Tib_all)
sheet.cell(2, 10).value = statistics.mean(R_Tib_all)
sheet.cell(3, 10).value = statistics.stdev(R_Tib_all)
# 腓腹筋内側頭
sheet.cell(2, 11).value = statistics.mean(L_Gas_all)
sheet.cell(3, 11).value = statistics.stdev(L_Gas_all)
sheet.cell(2, 12).value = statistics.mean(R_Gas_all)
sheet.cell(3, 12).value = statistics.stdev(R_Gas_all)

# 保存
book.save("excel/" + wbname)

print("終了")