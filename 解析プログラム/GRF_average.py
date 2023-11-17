

import os
import openpyxl
import pandas as pd
import numpy as np
import glob 
from tqdm import trange, tqdm
import statistics


#項目，相，条件入力
phase = "phase1"
condition_A = "condition2"
condition_B = "condition3"
condition_C = "condition4"
condition_D = "condition5"
condition_E = "condition6"
condition_F = "condition7"
condition_G = "condition8"
condition_H = "condition9"
condition_I = "condition10"

os.makedirs("analysis/GRF/" + phase + "/", exist_ok = True)

file_A = glob.glob("GRF/phase1/" + condition_A + "/*.csv")
file_B = glob.glob("GRF/phase1/" + condition_B + "/*.csv")
file_C = glob.glob("GRF/phase1/" + condition_C + "/*.csv")
file_D = glob.glob("GRF/phase1/" + condition_D + "/*.csv")
file_E = glob.glob("GRF/phase1/" + condition_E + "/*.csv")
file_F = glob.glob("GRF/phase1/" + condition_F + "/*.csv")
file_G = glob.glob("GRF/phase1/" + condition_G + "/*.csv")
file_H = glob.glob("GRF/phase1/" + condition_H + "/*.csv")
file_I = glob.glob("GRF/phase1/" + condition_I + "/*.csv")

list_A =[]
list_B =[]
list_C =[]
list_D =[]
list_E =[]
list_F =[]
list_G =[]
list_H =[]
list_I =[]

wbname = "GRF_average" + ".xlsx"
wb = openpyxl.Workbook()
wb.save("analysis/GRF/" + phase + "/" + wbname)

for file in tqdm(file_A):
    df = pd.read_csv(file)
    list = df["L_Sum"] + df["R_Sum"]

    list_A.extend(list)

for file in tqdm(file_B):
    df = pd.read_csv(file)
    list = df["L_Sum"] + df["R_Sum"]

    list_B.extend(list)

for file in tqdm(file_C):
    df = pd.read_csv(file)
    list = df["L_Sum"] + df["R_Sum"]

    list_C.extend(list)

for file in tqdm(file_D):
    df = pd.read_csv(file)
    list = df["L_Sum"] + df["R_Sum"]

    list_D.extend(list)

for file in tqdm(file_E):
    df = pd.read_csv(file)
    list = df["L_Sum"] + df["R_Sum"]

    list_E.extend(list)

for file in tqdm(file_F):
    df = pd.read_csv(file)
    list = df["L_Sum"] + df["R_Sum"]

    list_F.extend(list)

for file in tqdm(file_G):
    df = pd.read_csv(file)
    list = df["L_Sum"] + df["R_Sum"]

    list_G.extend(list)

for file in tqdm(file_H):
    df = pd.read_csv(file)
    list = df["L_Sum"] + df["R_Sum"]

    list_H.extend(list)

for file in tqdm(file_I):
    df = pd.read_csv(file)
    list = df["L_Sum"] + df["R_Sum"]

    list_I.extend(list)

book = openpyxl.load_workbook("analysis/GRF/" + phase + "/" + wbname)
sheet = book["Sheet"]

sheet.cell(1, 1).value = condition_A
sheet.cell(2, 1).value = statistics.mean(list_A)
sheet.cell(3, 1).value = statistics.stdev(list_A)
sheet.cell(1, 2).value = condition_B
sheet.cell(2, 2).value = statistics.mean(list_B)
sheet.cell(3, 2).value = statistics.stdev(list_B)
sheet.cell(1, 3).value = condition_C
sheet.cell(2, 3).value = statistics.mean(list_C)
sheet.cell(3, 3).value = statistics.stdev(list_C)
sheet.cell(1, 4).value = condition_D
sheet.cell(2, 4).value = statistics.mean(list_D)
sheet.cell(3, 4).value = statistics.stdev(list_D)
sheet.cell(1, 5).value = condition_E
sheet.cell(2, 5).value = statistics.mean(list_E)
sheet.cell(3, 5).value = statistics.stdev(list_E)
sheet.cell(1, 6).value = condition_F
sheet.cell(2, 6).value = statistics.mean(list_F)
sheet.cell(3, 6).value = statistics.stdev(list_F)
sheet.cell(1, 7).value = condition_G
sheet.cell(2, 7).value = statistics.mean(list_G)
sheet.cell(3, 7).value = statistics.stdev(list_G)
sheet.cell(1, 8).value = condition_H
sheet.cell(2, 8).value = statistics.mean(list_H)
sheet.cell(3, 8).value = statistics.stdev(list_H)
sheet.cell(1, 9).value = condition_I
sheet.cell(2, 9).value = statistics.mean(list_I)
sheet.cell(3, 9).value = statistics.stdev(list_I)

book.save("analysis/GRF/" + phase + "/" + wbname)