from statistics import mean
import pandas as pd
import numpy as np
import openpyxl
from tqdm import trange, tqdm
import os
import glob
from matplotlib import pyplot

filename = "o_"
phase = "phase1"
os.makedirs("analysis/GRF" + "/" + phase, exist_ok = True)

wbname = "GRF_mean_" + phase + ".xlsx"
wb = openpyxl.Workbook()
wb.save ("analysis/GRF/" + phase + "/GRF_mean_" + phase + ".xlsx")

book = openpyxl.load_workbook("analysis/GRF/" + phase + "/GRF_mean_" + phase + ".xlsx")
sheet =book["Sheet"]

for i in tqdm(range(1, 11)):
    for j in tqdm(range(1, 6)):
        try:
            df = pd.read_csv("phase/GRF/" + filename + str(i)  + "_" 
                                + str(j) + "_" + phase + ".csv")
                                
        
            Sum = df["L_Sum"] + df["R_Sum"]
            ave = Sum.mean() 

            sheet.cell(row = j, column = 2*i-1 ).value = filename + str(i) + "_" + str(j)
            sheet.cell(row = j, column = 2*i).value = ave
            sheet.cell(row = j, column = 2*i).number_format = "0.00"
        
            book.save("analysis/GRF/" + phase + "/GRF_mean_" + phase  + ".xlsx")
        except:
            print("")    


phase = "phase2"
os.makedirs("analysis/GRF" + "/" + phase, exist_ok = True)

wbname = "GRF_mean_" + phase + ".xlsx"
wb = openpyxl.Workbook()
wb.save ("analysis/GRF/" + phase + "/GRF_mean_" + phase + ".xlsx")

book = openpyxl.load_workbook("analysis/GRF/" + phase + "/GRF_mean_" + phase + ".xlsx")
sheet =book["Sheet"]

for i in tqdm(range(1, 11)):
    for j in tqdm(range(1, 6)):
        try:
            df = pd.read_csv("phase/GRF/" + filename + str(i)  + "_" 
                                    + str(j) + "_" + phase + ".csv")
                                    
            
            Sum = df["L_Sum"] + df["R_Sum"]
            ave = Sum.mean() 

            sheet.cell(row = j, column = 2*i-1 ).value = filename + str(i) + "_" + str(j)
            sheet.cell(row = j, column = 2*i).value = ave
            sheet.cell(row = j, column = 2*i).number_format = "0.00"
            
            book.save("analysis/GRF/" + phase + "/GRF_mean_" + phase  + ".xlsx")
        except:
            print("")    


phase = "phase3_1"
os.makedirs("analysis/GRF" + "/" + phase, exist_ok = True)

wbname = "GRF_mean_" + phase + ".xlsx"
wb = openpyxl.Workbook()
wb.save ("analysis/GRF/" + phase + "/GRF_mean_" + phase + ".xlsx")

book = openpyxl.load_workbook("analysis/GRF/" + phase + "/GRF_mean_" + phase + ".xlsx")
sheet =book["Sheet"]

for i in tqdm(range(1, 11)):
    for j in tqdm(range(1, 6)):
        try:
            df = pd.read_csv("phase/GRF/" + filename + str(i)  + "_" 
                                    + str(j) + "_" + phase + ".csv")
                                    
            
            Sum = df["L_Sum"] + df["R_Sum"]
            ave = Sum.mean() 

            sheet.cell(row = j, column = 2*i-1 ).value = filename + str(i) + "_" + str(j)
            sheet.cell(row = j, column = 2*i).value = ave
            sheet.cell(row = j, column = 2*i).number_format = "0.00"
            
            book.save("analysis/GRF/" + phase + "/GRF_mean_" + phase  + ".xlsx")
        except:
            print("")    


phase = "phase3_2"
os.makedirs("analysis/GRF" + "/" + phase, exist_ok = True)

wbname = "GRF_mean_" + phase + ".xlsx"
wb = openpyxl.Workbook()
wb.save ("analysis/GRF/" + phase + "/GRF_mean_" + phase + ".xlsx")

book = openpyxl.load_workbook("analysis/GRF/" + phase + "/GRF_mean_" + phase + ".xlsx")
sheet =book["Sheet"]

for i in tqdm(range(1, 11)):
    for j in tqdm(range(1, 6)):
        try:
            df = pd.read_csv("phase/GRF/" + filename + str(i)  + "_" 
                                    + str(j) + "_" + phase + ".csv")
                                    
            
            Sum = df["L_Sum"] + df["R_Sum"]
            ave = Sum.mean() 

            sheet.cell(row = j, column = 2*i-1 ).value = filename + str(i) + "_" + str(j)
            sheet.cell(row = j, column = 2*i).value = ave
            sheet.cell(row = j, column = 2*i).number_format = "0.00"
            
            book.save("analysis/GRF/" + phase + "/GRF_mean_" + phase  + ".xlsx")
        except:
            print("")    


phase = "phase3_3"
os.makedirs("analysis/GRF" + "/" + phase, exist_ok = True)

wbname = "GRF_mean_" + phase + ".xlsx"
wb = openpyxl.Workbook()
wb.save ("analysis/GRF/" + phase + "/GRF_mean_" + phase + ".xlsx")

book = openpyxl.load_workbook("analysis/GRF/" + phase + "/GRF_mean_" + phase + ".xlsx")
sheet =book["Sheet"]

for i in tqdm(range(1, 11)):
    for j in tqdm(range(1, 6)):
        try:
            df = pd.read_csv("phase/GRF/" + filename + str(i)  + "_" 
                                    + str(j) + "_" + phase + ".csv")
                                    
            
            Sum = df["L_Sum"] + df["R_Sum"]
            ave = Sum.mean() 

            sheet.cell(row = j, column = 2*i-1 ).value = filename + str(i) + "_" + str(j)
            sheet.cell(row = j, column = 2*i).value = ave
            sheet.cell(row = j, column = 2*i).number_format = "0.00"
            
            book.save("analysis/GRF/" + phase + "/GRF_mean_" + phase  + ".xlsx")
        except:
            print("")


phase = "phase3_4"
os.makedirs("analysis/GRF" + "/" + phase, exist_ok = True)

wbname = "GRF_mean_" + phase + ".xlsx"
wb = openpyxl.Workbook()
wb.save ("analysis/GRF/" + phase + "/GRF_mean_" + phase + ".xlsx")

book = openpyxl.load_workbook("analysis/GRF/" + phase + "/GRF_mean_" + phase + ".xlsx")
sheet =book["Sheet"]

for i in tqdm(range(1, 11)):
    for j in tqdm(range(1, 6)):
        try:
            df = pd.read_csv("phase/GRF/" + filename + str(i)  + "_" 
                                    + str(j) + "_" + phase + ".csv")
                                    
            
            Sum = df["L_Sum"] + df["R_Sum"]
            ave = Sum.mean() 

            sheet.cell(row = j, column = 2*i-1 ).value = filename + str(i) + "_" + str(j)
            sheet.cell(row = j, column = 2*i).value = ave
            sheet.cell(row = j, column = 2*i).number_format = "0.00"
            
            book.save("analysis/GRF/" + phase + "/GRF_mean_" + phase  + ".xlsx")
        except:
            print("")    