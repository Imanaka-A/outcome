#重心変位を時系列で表示
####被験者毎
import statistics
import pandas as pd
import glob
from tqdm import trange, tqdm
from natsort import natsorted
import openpyxl
import re
import os

#コピー用
CoM_LIST = ["CoM_x", "CoM_y", "CoM_z"]
PHASE_LIST = ["phase1", "phase2", "phase3_1", "phase3_2", "phase3_3", "phase3_4"]

NAME = ["i", "n", "o", "y"]

#処理項目
condition = "condition2"
phase = "phase1"

#リスト
list_0 = []
list_10 =[]
list_20 = []
list_30 = []
list_40 = []
list_50 = []
list_60 = []
list_70 = []
list_80 = []
list_90 = []
list_100 = []

#被験者ループ
for com in tqdm(CoM_LIST):
    print(com)

    #エクセル作成
    wbname = com + "_" + condition + "_" + phase + ".xlsx"
    wb = openpyxl.Workbook()
    wb.save("ex/" + wbname)

    for i, name in enumerate(NAME):
        print(name)

        #読み込み
        """
        file_10 = glob.glob(condition + "/" + phase + "/" + "EMG" + "/" + "10" + "/" + name + "/*.csv")
        file_20 = glob.glob(condition + "/" + phase + "/" + "EMG" + "/" + "20" + "/" + name + "/*.csv")
        file_30 = glob.glob(condition + "/" + phase + "/" + "EMG" + "/" + "30" + "/" + name + "/*.csv")
        file_40 = glob.glob(condition + "/" + phase + "/" + "EMG" + "/" + "40" + "/" + name + "/*.csv")
        file_50 = glob.glob(condition + "/" + phase + "/" + "EMG" + "/" + "50" + "/" + name + "/*.csv")
        file_60 = glob.glob(condition + "/" + phase + "/" + "EMG" + "/" + "60" + "/" + name + "/*.csv")
        file_70 = glob.glob(condition + "/" + phase + "/" + "EMG" + "/" + "70" + "/" + name + "/*.csv")
        file_80 = glob.glob(condition + "/" + phase + "/" + "EMG" + "/" + "80" + "/" + name + "/*.csv")
        file_90 = glob.glob(condition + "/" + phase + "/" + "EMG" + "/" + "90" + "/" + name + "/*.csv")
        file_100 = glob.glob(condition + "/" + phase + "/" + "EMG" + "/" + "100" + "/" + name + "/*.csv")
        """

        file_10 = glob.glob(condition + "/" + phase + "/" + "CoM_displacement" + "/" + "10/*.csv")
        file_20 = glob.glob(condition + "/" + phase + "/" + "CoM_displacement" + "/" + "20/*.csv")
        file_30 = glob.glob(condition + "/" + phase + "/" + "CoM_displacement" + "/" + "30/*.csv")
        file_40 = glob.glob(condition + "/" + phase + "/" + "CoM_displacement" + "/" + "40/*.csv")
        file_50 = glob.glob(condition + "/" + phase + "/" + "CoM_displacement" + "/" + "50/*.csv")
        file_60 = glob.glob(condition + "/" + phase + "/" + "CoM_displacement" + "/" + "60/*.csv")
        file_70 = glob.glob(condition + "/" + phase + "/" + "CoM_displacement" + "/" + "70/*.csv")
        file_80 = glob.glob(condition + "/" + phase + "/" + "CoM_displacement" + "/" + "80/*.csv")
        file_90 = glob.glob(condition + "/" + phase + "/" + "CoM_displacement" + "/" + "90/*.csv")
        file_100 = glob.glob(condition + "/" + phase + "/" + "CoM_displacement" + "/" + "100/*.csv")

        book = openpyxl.load_workbook("ex/" + wbname)

        #リストに追加
        for file in natsorted(file_10):
            df = pd.read_csv(file)
            splitname = os.path.basename(file).split("_") #ファイル名から被験者名取得
            if splitname[0] == name:
                try:
                    list = df[com]
                    list0 =df[com][0] #初期値
                except:
                    pass
                else:
                    list_0.append(list0)
                    list_10.extend(list)

        for file in natsorted(file_20):
            df = pd.read_csv(file)
            splitname = os.path.basename(file).split("_") #ファイル名から被験者名取得
            if splitname[0] == name:
                try:
                    list = df[com]
                except:
                    pass
                else:
                    list_20.extend(list)

        for file in natsorted(file_30):
            df = pd.read_csv(file)
            splitname = os.path.basename(file).split("_") #ファイル名から被験者名取得
            if splitname[0] == name:
                try:
                    list = df[com]
                except:
                    pass
                else:
                    list_30.extend(list)

        for file in natsorted(file_40):
            df = pd.read_csv(file)
            splitname = os.path.basename(file).split("_") #ファイル名から被験者名取得
            if splitname[0] == name:
                try:
                    list = df[com]
                except:
                    pass
                else:
                    list_40.extend(list)

        for file in natsorted(file_50):
            df = pd.read_csv(file)
            splitname = os.path.basename(file).split("_") #ファイル名から被験者名取得
            if splitname[0] == name:
                try:
                    list = df[com]
                except:
                    pass
                else:
                    list_50.extend(list)

        for file in natsorted(file_60):
            df = pd.read_csv(file)
            splitname = os.path.basename(file).split("_") #ファイル名から被験者名取得
            if splitname[0] == name:
                try:
                    list = df[com]
                except:
                    pass
                else:
                    list_60.extend(list)

        for file in natsorted(file_70):
            df = pd.read_csv(file)
            splitname = os.path.basename(file).split("_") #ファイル名から被験者名取得
            if splitname[0] == name:
                try:
                    list = df[com]
                except:
                    pass
                else:
                    list_70.extend(list)

        for file in natsorted(file_80):
            df = pd.read_csv(file)
            splitname = os.path.basename(file).split("_") #ファイル名から被験者名取得
            if splitname[0] == name:
                try:
                    list = df[com]
                except:
                    pass
                else:
                    list_80.extend(list)

        for file in natsorted(file_90):
            df = pd.read_csv(file)
            splitname = os.path.basename(file).split("_") #ファイル名から被験者名取得
            if splitname[0] == name:
                try:
                    list = df[com]
                except:
                    pass
                else:
                    list_90.extend(list)

        for file in natsorted(file_100):
            df = pd.read_csv(file)
            splitname = os.path.basename(file).split("_") #ファイル名から被験者名取得
            if splitname[0] == name:
                try:
                    list = df[com]
                except:
                    pass
                else:
                    list_100.extend(list)

        sheet = book.worksheets[i]
        sheet.title = name

        #割合記入
        for j in range(0, 11):
            time_list = [0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100 ]
            sheet.cell(j + 1, 1).value =time_list[j]

        #平均値と標準偏差を記入
        sheet.cell(1, 2).value = statistics.mean(list_0)
        sheet.cell(1, 3).value = statistics.stdev(list_0)
        sheet.cell(2, 2).value = statistics.mean(list_10)
        sheet.cell(2, 3).value = statistics.stdev(list_10)
        sheet.cell(3, 2).value = statistics.mean(list_20)
        sheet.cell(3, 3).value = statistics.stdev(list_20)
        sheet.cell(4, 2).value = statistics.mean(list_30)
        sheet.cell(4, 3).value = statistics.stdev(list_30)
        sheet.cell(5, 2).value = statistics.mean(list_40)
        sheet.cell(5, 3).value = statistics.stdev(list_40)
        sheet.cell(6, 2).value = statistics.mean(list_50)
        sheet.cell(6, 3).value = statistics.stdev(list_50)
        sheet.cell(7, 2).value = statistics.mean(list_60)
        sheet.cell(7, 3).value = statistics.stdev(list_60)
        sheet.cell(8, 2).value = statistics.mean(list_70)
        sheet.cell(8, 3).value = statistics.stdev(list_70)
        sheet.cell(9, 2).value = statistics.mean(list_80)
        sheet.cell(9, 3).value = statistics.stdev(list_80)
        sheet.cell(10, 2).value = statistics.mean(list_90)
        sheet.cell(10, 3).value = statistics.stdev(list_90)
        sheet.cell(11, 2).value = statistics.mean(list_100)
        sheet.cell(11, 3).value = statistics.stdev(list_100)

        book.create_sheet()
        book.save("ex/" + wbname)

        #リストをリセット
        list_0.clear()
        list_10.clear()
        list_20.clear()
        list_30.clear()
        list_40.clear()
        list_50.clear()
        list_60.clear()
        list_70.clear()
        list_80.clear()
        list_90.clear()
        list_100.clear()

print("終了")