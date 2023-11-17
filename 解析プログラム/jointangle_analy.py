#関節角度を%ごとにまとめてエクセル表示
#相，条件を全て処理可能

import pandas as pd
import openpyxl
from tqdm import trange, tqdm
import os
import glob
import statistics

#for文用リスト
##条件数に応じてconditionは変更
phases = ["phase1", "phase2"]
conditions = ["condition2", "condition3", "condition4","condition5",
                "condition6", "condition7", "condition8", "condition9", "condition10"]

for condition in tqdm(conditions):

    for phase in phases:

        #ファイル作成
        os.makedirs("analysis/angle/crotch/" + phase, exist_ok = True)
        #空リスト
        LIST_0 = []
        LIST_1 = []
        LIST_2 = []
        LIST_3 = []
        LIST_4 = []
        LIST_5 = []
        LIST_6 = []
        LIST_7 = []
        LIST_8 = []
        LIST_9 = []
        LIST_10 = []

        #分割したファイル読み込み
        files_1 = glob.glob(condition + "/" + phase + "/angle/10/*.csv")
        files_2 = glob.glob(condition + "/" + phase + "/angle/20/*.csv")
        files_3 = glob.glob(condition + "/" + phase + "/angle/30/*.csv")
        files_4 = glob.glob(condition + "/" + phase + "/angle/40/*.csv")
        files_5 = glob.glob(condition + "/" + phase + "/angle/50/*.csv")
        files_6 = glob.glob(condition + "/" + phase + "/angle/60/*.csv")
        files_7 = glob.glob(condition + "/" + phase + "/angle/70/*.csv")
        files_8 = glob.glob(condition + "/" + phase + "/angle/80/*.csv")
        files_9 = glob.glob(condition + "/" + phase + "/angle/90/*.csv")
        files_10 = glob.glob(condition + "/" + phase + "/angle/100/*.csv")

        #エクセル作成
        #要変更
        wbname  ="Rcrotch" + "_" + phase + "_" + condition + ".xlsx"
        wb = openpyxl.Workbook()
        wb.save("analysis/angle/crotch/" + phase + "/" + wbname)

        book = openpyxl.load_workbook("analysis/angle/crotch/" + phase + "/" + wbname)
        sheet = book["Sheet"]

        #%ごとに処理
        #条件1
        for file in files_1:
            df = pd.read_csv(file)
            LIST = df["R_crotch_degree_xy"]
            LIST0 =df["R_crotch_degree_xy"][0]

            LIST_1.extend(LIST)
            LIST_0.append(LIST0)

        for i in range(0, len(LIST_1)):

            sheet.cell(i + 4, 1).value = LIST_1[i]

        book.save("analysis/angle/crotch/" + phase + "/" + wbname)

        #条件2
        for file in files_2:
            df = pd.read_csv(file)
            LIST = df["R_crotch_degree_xy"]

            LIST_2.extend(LIST)

        for i in range(0, len(LIST_2)):

            sheet.cell(i + 4, 2).value = LIST_2[i]

        book.save("analysis/angle/crotch/" + phase + "/" + wbname)

        #条件3
        for file in files_3:
            df = pd.read_csv(file)
            LIST = df["R_crotch_degree_xy"]

            LIST_3.extend(LIST)

        for i in range(0, len(LIST_3)):

            sheet.cell(i + 4, 3).value = LIST_3[i]

        book.save("analysis/angle/crotch/" + phase + "/" + wbname)

        #条件4
        for file in files_4:
            df = pd.read_csv(file)
            LIST = df["R_crotch_degree_xy"]

            LIST_4.extend(LIST)

        for i in range(0, len(LIST_4)):

            sheet.cell(i + 4, 4).value = LIST_4[i]

        book.save("analysis/angle/crotch/" + phase + "/" + wbname)

        #条件5
        for file in files_5:
            df = pd.read_csv(file)
            LIST = df["R_crotch_degree_xy"]

            LIST_5.extend(LIST)

        for i in range(0, len(LIST_5)):

            sheet.cell(i + 4, 5).value = LIST_5[i]

        book.save("analysis/angle/crotch/" + phase + "/" + wbname)

        #条件6
        for file in files_6:
            df = pd.read_csv(file)
            LIST = df["R_crotch_degree_xy"]

            LIST_6.extend(LIST)

        for i in range(0, len(LIST_6)):

            sheet.cell(i + 4, 6).value = LIST_6[i]

        book.save("analysis/angle/crotch/" + phase + "/" + wbname)

        #条件7
        for file in files_7:
            df = pd.read_csv(file)
            LIST = df["R_crotch_degree_xy"]

            LIST_7.extend(LIST)

        for i in range(0, len(LIST_7)):

            sheet.cell(i + 4, 7).value = LIST_7[i]

        book.save("analysis/angle/crotch/" + phase + "/" + wbname)

        #条件8
        for file in files_8:
            df = pd.read_csv(file)
            LIST = df["R_crotch_degree_xy"]

            LIST_8.extend(LIST)

        for i in range(0, len(LIST_8)):

            sheet.cell(i + 4, 8).value = LIST_8[i]

        book.save("analysis/angle/crotch/" + phase + "/" + wbname)

        #条件9
        for file in files_9:
            df = pd.read_csv(file)
            LIST = df["R_crotch_degree_xy"]

            LIST_9.extend(LIST)

        for i in range(0, len(LIST_9)):

            sheet.cell(i + 4, 9).value = LIST_9[i]

        book.save("analysis/angle/crotch/" + phase + "/" + wbname)

        #条件10
        for file in files_10:
            df = pd.read_csv(file)
            LIST = df["R_crotch_degree_xy"]

            LIST_10.extend(LIST)

        for i in range(0, len(LIST_10)):

            sheet.cell(i + 4, 10).value = LIST_10[i]

        book.save("analysis/angle/crotch/" + phase + "/" + wbname)

        ##まとめ用エクセル作成
        wbname2 = wbname.replace(".xlsx","") + "_まとめ" + ".xlsx"
        wb = openpyxl.Workbook()
        wb.save("analysis/angle/crotch/" + phase + "/" + wbname2)

        book = openpyxl.load_workbook("analysis/angle/crotch/" + phase + "/" + wbname2)
        sheet = book["Sheet"]

        for i in range(0, 11):
            excel_list = [0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100 ]
            sheet.cell(i + 3, 2).value = excel_list[i]

        book.save("analysis/angle/crotch/" + phase + "/" + wbname2)

        sheet.cell(3, 3).value =statistics.mean(LIST_0)
        sheet.cell(3, 4).value =statistics.stdev(LIST_0)
        sheet.cell(4, 3).value = statistics.mean(LIST_1)
        sheet.cell(4, 4).value = statistics.stdev(LIST_1)
        sheet.cell(5, 3).value = statistics.mean(LIST_2)
        sheet.cell(5, 4).value = statistics.stdev(LIST_2)
        sheet.cell(6, 3).value = statistics.mean(LIST_3)
        sheet.cell(6, 4).value = statistics.stdev(LIST_3)
        sheet.cell(7, 3).value = statistics.mean(LIST_4)
        sheet.cell(7, 4).value = statistics.stdev(LIST_4)
        sheet.cell(8, 3).value = statistics.mean(LIST_5)
        sheet.cell(8, 4).value = statistics.stdev(LIST_5)
        sheet.cell(9, 3).value = statistics.mean(LIST_6)
        sheet.cell(9, 4).value = statistics.stdev(LIST_6)
        sheet.cell(10, 3).value = statistics.mean(LIST_7)
        sheet.cell(10, 4).value = statistics.stdev(LIST_7)
        sheet.cell(11, 3).value = statistics.mean(LIST_8)
        sheet.cell(11, 4).value = statistics.stdev(LIST_8)
        sheet.cell(12, 3).value = statistics.mean(LIST_9)
        sheet.cell(12, 4).value = statistics.stdev(LIST_9)
        sheet.cell(13, 3).value = statistics.mean(LIST_10)
        sheet.cell(13, 4).value = statistics.stdev(LIST_10)

        book.save("analysis/angle/crotch/" + phase + "/" + wbname2)

        #リストをリセット
        LIST_0.clear()
        LIST_1.clear()
        LIST_2.clear()
        LIST_3.clear()
        LIST_4.clear()
        LIST_5.clear()
        LIST_6.clear()
        LIST_7.clear()
        LIST_8.clear()
        LIST_9.clear()
        LIST_10.clear()

print("終了")