#%ごとにまとめてエクセル表示
#相，条件を全て処理可能

import pandas as pd
import openpyxl
from tqdm import trange, tqdm
import os
import glob
import statistics

#for文用リスト
##条件数に応じてconditionは変更
phases = ["phase1", "phase2", "phase3_1", "phase3_2", "phase3_3", "phase3_4"]
conditions = ["condition1", "condition2", "condition3", "condition4","condition5",
                "condition6", "condition7", "condition8", "condition9", "condition10"]

for condition in conditions:
    
    for phase in phases:

        os.makedirs("analysis/GRF/" + phase, exist_ok = True)
        
        #空リスト
        list_0 = []
        list_1 = []
        list_2 = []
        list_3 = []
        list_4 = []
        list_5 = []
        list_6 = []
        list_7 = []
        list_8 = []
        list_9 = []
        list_10 = []
        
        #分割したファイル読み込み
        files_1 = glob.glob(condition + "/" + phase + "/GRF/10/*.csv")
        files_2 = glob.glob(condition + "/" + phase + "/GRF/20/*.csv")
        files_3 = glob.glob(condition + "/" + phase + "/GRF/30/*.csv")
        files_4 = glob.glob(condition + "/" + phase + "/GRF/40/*.csv")
        files_5 = glob.glob(condition + "/" + phase + "/GRF/50/*.csv")
        files_6 = glob.glob(condition + "/" + phase + "/GRF/60/*.csv")
        files_7 = glob.glob(condition + "/" + phase + "/GRF/70/*.csv")
        files_8 = glob.glob(condition + "/" + phase + "/GRF/80/*.csv")
        files_9 = glob.glob(condition + "/" + phase + "/GRF/90/*.csv")
        files_10 = glob.glob(condition + "/" + phase + "/GRF/100/*.csv")

        #エクセル作成
        wbname  ="GRF" + "_" + phase + "_" + condition + ".xlsx"
        wb = openpyxl.Workbook()
        wb.save("analysis/GRF/" + phase + "/" + wbname)

        book = openpyxl.load_workbook("analysis/GRF/" + phase + "/" + wbname)
        sheet = book["Sheet"]

        #%ごとに処理
        #10
        for file in tqdm(files_1):
            df = pd.read_csv(file)
            list = df["R_Sum"] + df["L_Sum"]
            list0 = df["R_Sum"][0] + df["L_Sum"][0]

            list_1.extend(list)
            list_0.append(list0)       

        for i in range(0, len(list_1)):
            
            sheet.cell(i + 4, 1).value = list_1[i]

        book.save("analysis/GRF/" + phase + "/" + wbname)

        #20
        for file in tqdm(files_2):
            df = pd.read_csv(file)
            list = df["R_Sum"] + df["L_Sum"]

            list_2.extend(list)

        for i in range(0, len(list_2)):
            
            sheet.cell(i + 4, 2).value = list_2[i]

        book.save("analysis/GRF/" + phase + "/" + wbname)

        #30
        for file in tqdm(files_3):
            df = pd.read_csv(file)
            list = df["R_Sum"] + df["L_Sum"]

            list_3.extend(list)

        for i in range(0, len(list_3)):
            
            sheet.cell(i + 4, 3).value = list_3[i]

        book.save("analysis/GRF/" + phase + "/" + wbname)

        #40
        for file in tqdm(files_4):
            df = pd.read_csv(file)
            list = df["R_Sum"] +df["L_Sum"]

            list_4.extend(list)

        for i in range(0, len(list_4)):
            
            sheet.cell(i + 4, 4).value = list_4[i]

        book.save("analysis/GRF/" + phase + "/" + wbname)

        #50
        for file in tqdm(files_5):
            df = pd.read_csv(file)
            list = df["R_Sum"] + df["L_Sum"]

            list_5.extend(list)

        for i in range(0, len(list_5)):
            
            sheet.cell(i + 4, 5).value = list_5[i]

        book.save("analysis/GRF/" + phase + "/" + wbname)

        #60
        for file in tqdm(files_6):
            df = pd.read_csv(file)
            list = df["R_Sum"] + df["L_Sum"]

            list_6.extend(list)

        for i in range(0, len(list_6)):
            
            sheet.cell(i + 4, 6).value = list_6[i]

        book.save("analysis/GRF/" + phase + "/" + wbname)

        #70
        for file in tqdm(files_7):
            df = pd.read_csv(file)
            list = df["R_Sum"] + df["L_Sum"]

            list_7.extend(list)

        for i in range(0, len(list_7)):
            
            sheet.cell(i + 4, 7).value = list_7[i]

        book.save("analysis/GRF/" + phase + "/" + wbname)

        #80
        for file in tqdm(files_8):
            df = pd.read_csv(file)
            list = df["R_Sum"] + df["L_Sum"]

            list_8.extend(list)

        for i in range(0, len(list_8)):
            
            sheet.cell(i + 4, 8).value = list_8[i]

        book.save("analysis/GRF/" + phase + "/" + wbname)

        #90
        for file in tqdm(files_9):
            df = pd.read_csv(file)
            list = df["R_Sum"] + df["L_Sum"]

            list_9.extend(list)

        for i in range(0, len(list_9)):
            
            sheet.cell(i + 4, 9).value = list_9[i]

        book.save("analysis/GRF/" + phase + "/" + wbname)

        #100
        for file in tqdm(files_10):
            df = pd.read_csv(file)
            list = df["R_Sum"] + df["L_Sum"]

            list_10.extend(list)

        for i in range(0, len(list_10)):
            
            sheet.cell(i + 4, 10).value = list_10[i]

        book.save("analysis/GRF/" + phase + "/" + wbname)

        ##まとめ用エクセル作成
        wbname = "GRF" + "_" + phase + "_" + condition + "_まとめ" + ".xlsx"
        wb = openpyxl.Workbook()
        wb.save("analysis/GRF/" + phase + "/" + wbname)

        book = openpyxl.load_workbook("analysis/GRF/" + phase + "/" + wbname)
        sheet = book["Sheet"]

        for i in range(0, 11):
            excel_list = [0, 10, 20, 30, 40, 50, 60, 70, 80, 90, 100 ]
            sheet.cell(i + 3, 2).value = excel_list[i]

        book.save("analysis/GRF/" + phase + "/" + wbname) 

        sheet.cell(3, 3).value = statistics.mean(list_0)
        sheet.cell(3, 4).value = statistics.stdev(list_0)
        sheet.cell(4, 3).value = statistics.mean(list_1)
        sheet.cell(4, 4).value = statistics.stdev(list_1)
        sheet.cell(5, 3).value = statistics.mean(list_2)
        sheet.cell(5, 4).value = statistics.stdev(list_2)
        sheet.cell(6, 3).value = statistics.mean(list_3)
        sheet.cell(6, 4).value = statistics.stdev(list_3)
        sheet.cell(7, 3).value = statistics.mean(list_4)
        sheet.cell(7, 4).value = statistics.stdev(list_4)
        sheet.cell(8, 3).value = statistics.mean(list_5)
        sheet.cell(8, 4).value = statistics.stdev(list_5)
        sheet.cell(9, 3).value = statistics.mean(list_6)
        sheet.cell(9, 4).value = statistics.stdev(list_6)
        sheet.cell(10, 3).value = statistics.mean(list_7)
        sheet.cell(10, 4).value = statistics.stdev(list_7)
        sheet.cell(11, 3).value = statistics.mean(list_8)
        sheet.cell(11, 4).value = statistics.stdev(list_8)
        sheet.cell(12, 3).value = statistics.mean(list_9)
        sheet.cell(12, 4).value = statistics.stdev(list_9)
        sheet.cell(13, 3).value = statistics.mean(list_10)
        sheet.cell(13, 4).value = statistics.stdev(list_10)

        book.save("analysis/GRF/" + phase + "/" + wbname)
                
        #リストをリセット
        list_0.clear()
        list_1.clear()
        list_2.clear()
        list_3.clear()
        list_4.clear()
        list_5.clear()
        list_6.clear()
        list_7.clear()
        list_8.clear()
        list_9.clear()
        list_10.clear()