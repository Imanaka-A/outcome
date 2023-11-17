#重心を%ごとにまとめてエクセル表示

import pandas as pd
import openpyxl
from tqdm import trange, tqdm
import os
import glob

#空リスト

phases = ["phase1", "phase2", "phase3_1", "phase3_2", "phase3_3", "phase3_4"]
conditions = ["condition1", "condition2", "condition3", "condition4","condition5",
                "condition6", "condition7", "condition8", "condition9", "condition10"]

for condition in conditions:
    
    for phase in phases:

        os.makedirs("analysis/CoM_y/" + phase, exist_ok = True)
        CoM_y_1 = []
        CoM_y_2 = []
        CoM_y_3 = []
        CoM_y_4 = []
        CoM_y_5 = []
        CoM_y_6 = []
        CoM_y_7 = []
        CoM_y_8 = []
        CoM_y_9 = []
        CoM_y_10 = []

        #分割したファイル読み込み
        files_1 = glob.glob(condition + "/" + phase + "/CoM/10/*.csv")
        files_2 = glob.glob(condition + "/" + phase + "/CoM/20/*.csv")
        files_3 = glob.glob(condition + "/" + phase + "/CoM/30/*.csv")
        files_4 = glob.glob(condition + "/" + phase + "/CoM/40/*.csv")
        files_5 = glob.glob(condition + "/" + phase + "/CoM/50/*.csv")
        files_6 = glob.glob(condition + "/" + phase + "/CoM/60/*.csv")
        files_7 = glob.glob(condition + "/" + phase + "/CoM/70/*.csv")
        files_8 = glob.glob(condition + "/" + phase + "/CoM/80/*.csv")
        files_9 = glob.glob(condition + "/" + phase + "/CoM/90/*.csv")
        files_10 = glob.glob(condition + "/" + phase + "/CoM/100/*.csv")

        #エクセル作成
        wbname  ="CoM_y_" + phase + "_" + condition + ".xlsx"
        wb = openpyxl.Workbook()
        wb.save("analysis/CoM_y/" + phase + "/CoM_y_" + phase + "_" + condition + ".xlsx")

        book = openpyxl.load_workbook("analysis/CoM_y/" + phase + "/CoM_y_" + phase + "_" + condition + ".xlsx")
        sheet = book["Sheet"]

        #%ごとに処理
        for file in tqdm(files_1):
            df = pd.read_csv(file)
            CoM_y = df["CoM_y"]

            CoM_y_1.extend(CoM_y)

        for i in range(0, len(CoM_y_1)):
            
            sheet.cell(i + 4, 1).value = CoM_y_1[i]

        book.save("analysis/CoM_y/" + phase + "/CoM_y_" + phase + "_" + condition + ".xlsx")

        for file in tqdm(files_2):
            df = pd.read_csv(file)
            CoM_y = df["CoM_y"]

            CoM_y_2.extend(CoM_y)

        for i in range(0, len(CoM_y_2)):
            
            sheet.cell(i + 4, 2).value = CoM_y_2[i]

        book.save("analysis/CoM_y/" + phase + "/CoM_y_" + phase + "_" + condition + ".xlsx")

        for file in tqdm(files_3):
            df = pd.read_csv(file)
            CoM_y = df["CoM_y"]

            CoM_y_3.extend(CoM_y)

        for i in range(0, len(CoM_y_3)):
            
            sheet.cell(i + 4, 3).value = CoM_y_3[i]

        book.save("analysis/CoM_y/" + phase + "/CoM_y_" + phase + "_" + condition + ".xlsx")

        for file in tqdm(files_4):
            df = pd.read_csv(file)
            CoM_y = df["CoM_y"]

            CoM_y_4.extend(CoM_y)

        for i in range(0, len(CoM_y_4)):
            
            sheet.cell(i + 4, 4).value = CoM_y_4[i]

        book.save("analysis/CoM_y/" + phase + "/CoM_y_" + phase + "_" + condition + ".xlsx")

        for file in tqdm(files_5):
            df = pd.read_csv(file)
            CoM_y = df["CoM_y"]

            CoM_y_5.extend(CoM_y)

        for i in range(0, len(CoM_y_5)):
            
            sheet.cell(i + 4, 5).value = CoM_y_5[i]

        book.save("analysis/CoM_y/" + phase + "/CoM_y_" + phase + "_" + condition + ".xlsx")

        for file in tqdm(files_6):
            df = pd.read_csv(file)
            CoM_y = df["CoM_y"]

            CoM_y_6.extend(CoM_y)

        for i in range(0, len(CoM_y_6)):
            
            sheet.cell(i + 4, 6).value = CoM_y_6[i]

        book.save("analysis/CoM_y/" + phase + "/CoM_y_" + phase + "_" + condition + ".xlsx")

        for file in tqdm(files_7):
            df = pd.read_csv(file)
            CoM_y = df["CoM_y"]

            CoM_y_7.extend(CoM_y)

        for i in range(0, len(CoM_y_7)):
            
            sheet.cell(i + 4, 7).value = CoM_y_7[i]

        book.save("analysis/CoM_y/" + phase + "/CoM_y_" + phase + "_" + condition + ".xlsx")

        for file in tqdm(files_8):
            df = pd.read_csv(file)
            CoM_y = df["CoM_y"]

            CoM_y_8.extend(CoM_y)

        for i in range(0, len(CoM_y_8)):
            
            sheet.cell(i + 4, 8).value = CoM_y_8[i]

        book.save("analysis/CoM_y/" + phase + "/CoM_y_" + phase + "_" + condition + ".xlsx")

        for file in tqdm(files_9):
            df = pd.read_csv(file)
            CoM_y = df["CoM_y"]

            CoM_y_9.extend(CoM_y)

        for i in range(0, len(CoM_y_9)):
            
            sheet.cell(i + 4, 9).value = CoM_y_9[i]

        book.save("analysis/CoM_y/" + phase + "/CoM_y_" + phase + "_" + condition + ".xlsx")

        for file in tqdm(files_10):
            df = pd.read_csv(file)
            CoM_y = df["CoM_y"]

            CoM_y_10.extend(CoM_y)

        for i in range(0, len(CoM_y_10)):
            
            sheet.cell(i + 4, 10).value = CoM_y_10[i]

        book.save("analysis/CoM_y/" + phase + "/CoM_y_" + phase + "_" + condition + ".xlsx")

        CoM_y_1.clear()
        CoM_y_2.clear()
        CoM_y_3.clear()
        CoM_y_4.clear()
        CoM_y_5.clear()
        CoM_y_6.clear()
        CoM_y_7.clear()
        CoM_y_8.clear()
        CoM_y_9.clear()
        CoM_y_10.clear()