import numpy as np
import pandas as pd
import glob
from tqdm import trange, tqdm
import os
from natsort import natsorted
from Tukey_function2 import tukey_hsd
import openpyxl

os.makedirs("Excel_EMG_Tukey/phase1/", exist_ok = True)
os.makedirs("Excel_EMG_Tukey/phase2/", exist_ok = True)
os.makedirs("Excel_EMG_Tukey/phase3_1/", exist_ok = True)
os.makedirs("Excel_EMG_Tukey/phase3_2/", exist_ok = True)
os.makedirs("Excel_EMG_Tukey/phase3_3/", exist_ok = True)
os.makedirs("Excel_EMG_Tukey/phase3_4/", exist_ok = True)

phase_list = ["phase1", "phase2", "phase3_1", "phase3_2", "phase3_3", "phase3_4"]

#処理したい条件
conditionA = "condition2"
conditionB = "condition3"
conditionC = "condition4"

#データ個数
data_number = 4

for phase in tqdm(phase_list):

    print("・・・" + phase)

    phase_number = phase

    #Excelファイル名
    filename = phase_number + "_" + conditionA + conditionB + conditionC

    #条件ごとにファイル
    file_A = glob.glob(conditionA + "/" + "phase/" + phase_number + "/*.csv")
    file_B = glob.glob(conditionB + "/" + "phase/" + phase_number + "/*.csv")
    file_C = glob.glob(conditionC + "/" + "phase/" + phase_number + "/*.csv")

    #空リスト
    #conditionA
    L_Bic_Br_A = []
    R_Bic_Br_A = []
    L_Med_tri_A = []
    R_Med_tri_A = []
    L_Rec_Fem_A = []
    R_Rec_Fem_A = []
    L_Vlo_A = []
    R_Vlo_A = []
    L_Tib_Ant_A = []
    R_Tib_Ant_A = []
    L_Med_Gas_A = []
    R_Med_Gas_A = []
    #conditionB
    L_Bic_Br_B = []
    R_Bic_Br_B = []
    L_Med_tri_B = []
    R_Med_tri_B = []
    L_Rec_Fem_B = []
    R_Rec_Fem_B = []
    L_Vlo_B = []
    R_Vlo_B = []
    L_Tib_Ant_B = []
    R_Tib_Ant_B = []
    L_Med_Gas_B = []
    R_Med_Gas_B = []
    #conditionC
    L_Bic_Br_C = []
    R_Bic_Br_C = []
    L_Med_tri_C = []
    R_Med_tri_C = []
    L_Rec_Fem_C = []
    R_Rec_Fem_C = []
    L_Vlo_C = []
    R_Vlo_C = []
    L_Tib_Ant_C = []
    R_Tib_Ant_C = []
    L_Med_Gas_C = []
    R_Med_Gas_C = []

    print(conditionA)
    for file in tqdm(natsorted(file_A)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_A.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_A.append(R_Bic_Br)
        try:
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_A.append(L_Med_tri)
        try:
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_A.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_A.append(L_Rec_Fem)
        try:
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_A.append(R_Rec_Fem)
        try:
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_A.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass
        else:
            R_Vlo_A.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
            L_Tib_Ant_A.append(L_Tib_Ant)
        try:
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
            R_Tib_Ant_A.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_A.append(L_Med_Gas)
        try:
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            R_Med_Gas_A.append(R_Med_Gas)

    #配列に変換
    L_Bic_Br_A = pd.Series(L_Bic_Br_A)
    R_Bic_Br_A = pd.Series(R_Bic_Br_A)
    L_Med_tri_A = pd.Series(L_Med_tri_A)
    R_Med_tri_A = pd.Series(R_Med_tri_A)
    L_Rec_Fem_A = pd.Series(L_Rec_Fem_A)
    R_Rec_Fem_A = pd.Series(R_Rec_Fem_A)
    L_Vlo_A = pd.Series(L_Vlo_A)
    R_Vlo_A = pd.Series(R_Vlo_A)
    L_Tib_Ant_A = pd.Series(L_Tib_Ant_A)
    R_Tib_Ant_A = pd.Series(R_Tib_Ant_A)
    L_Med_Gas_A = pd.Series(L_Med_Gas_A)
    R_Med_Gas_A = pd.Series(R_Med_Gas_A)

    print(conditionB)
    for file in tqdm(natsorted(file_B)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_B.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_B.append(R_Bic_Br)
        try:
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_B.append(L_Med_tri)
        try:
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_B.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_B.append(L_Rec_Fem)
        try:
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_B.append(R_Rec_Fem)
        try:
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_B.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass
        else:
            R_Vlo_B.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
            L_Tib_Ant_B.append(L_Tib_Ant)
        try:
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
            R_Tib_Ant_B.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_B.append(L_Med_Gas)
        try:
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            R_Med_Gas_B.append(R_Med_Gas)

    #配列に変換
    L_Bic_Br_B = pd.Series(L_Bic_Br_B)
    R_Bic_Br_B = pd.Series(R_Bic_Br_B)
    L_Med_tri_B = pd.Series(L_Med_tri_B)
    R_Med_tri_B = pd.Series(R_Med_tri_B)
    L_Rec_Fem_B = pd.Series(L_Rec_Fem_B)
    R_Rec_Fem_B = pd.Series(R_Rec_Fem_B)
    L_Vlo_B = pd.Series(L_Vlo_B)
    R_Vlo_B = pd.Series(R_Vlo_B)
    L_Tib_Ant_B = pd.Series(L_Tib_Ant_B)
    R_Tib_Ant_B = pd.Series(R_Tib_Ant_B)
    L_Med_Gas_B = pd.Series(L_Med_Gas_B)
    R_Med_Gas_B = pd.Series(R_Med_Gas_B)

    print(conditionC)
    for file in tqdm(natsorted(file_C)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_C.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_C.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_C.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_C.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_C.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_C.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_C.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_C.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_C.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
             R_Tib_Ant_C.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_C.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_C.append(R_Med_Gas)

    #配列に変換
    L_Bic_Br_C = pd.Series(L_Bic_Br_C)
    R_Bic_Br_C = pd.Series(R_Bic_Br_C)
    L_Med_tri_C = pd.Series(L_Med_tri_C)
    R_Med_tri_C = pd.Series(R_Med_tri_C)
    L_Rec_Fem_C = pd.Series(L_Rec_Fem_C)
    R_Rec_Fem_C = pd.Series(R_Rec_Fem_C)
    L_Vlo_C = pd.Series(L_Vlo_C)
    R_Vlo_C = pd.Series(R_Vlo_C)
    L_Tib_Ant_C = pd.Series(L_Tib_Ant_C)
    R_Tib_Ant_C = pd.Series(R_Tib_Ant_C)
    L_Med_Gas_C = pd.Series(L_Med_Gas_C)
    R_Med_Gas_C = pd.Series(R_Med_Gas_C)

    #検定スタート
    #左上腕二頭筋
    A = L_Bic_Br_A
    B = L_Bic_Br_B
    C = L_Bic_Br_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Bic_Br = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Bic_Br = pd.concat([df_L_Bic_Br, df_condition, df_condition.describe()], axis = 1)

    #右上腕二頭筋
    A = R_Bic_Br_A
    B = R_Bic_Br_B
    C = R_Bic_Br_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Bic_Br = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Bic_Br = pd.concat([df_R_Bic_Br, df_condition, df_condition.describe()], axis = 1)

    #左上腕三頭筋
    A = L_Med_tri_A
    B = L_Med_tri_B
    C = L_Med_tri_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Med_tri = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Med_tri = pd.concat([df_L_Med_tri, df_condition, df_condition.describe()], axis = 1)

    #右上腕三頭筋
    A = R_Med_tri_A
    B = R_Med_tri_B
    C = R_Med_tri_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Med_tri = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Med_tri = pd.concat([df_R_Med_tri, df_condition, df_condition.describe()], axis = 1)

    #左大腿直筋
    A = L_Rec_Fem_A
    B = L_Rec_Fem_B
    C = L_Rec_Fem_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Rec_Fem = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Rec_Fem = pd.concat([df_L_Rec_Fem, df_condition, df_condition.describe()], axis = 1)

    #右大腿直筋
    A = R_Rec_Fem_A
    B = R_Rec_Fem_B
    C = R_Rec_Fem_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Rec_Fem = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Rec_Fem = pd.concat([df_R_Rec_Fem, df_condition, df_condition.describe()], axis = 1)

    #左外側広筋
    A = L_Vlo_A
    B = L_Vlo_B
    C = L_Vlo_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Vlo = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Vlo = pd.concat([df_L_Vlo, df_condition, df_condition.describe()], axis = 1)

    #右外側広筋
    A = R_Vlo_A
    B = R_Vlo_B
    C = R_Vlo_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Vlo = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Vlo = pd.concat([df_R_Vlo, df_condition, df_condition.describe()], axis = 1)

    #左前脛骨筋
    A = L_Tib_Ant_A
    B = L_Tib_Ant_B
    C = L_Tib_Ant_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Tib_Ant = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Tib_Ant = pd.concat([df_L_Tib_Ant, df_condition, df_condition.describe()], axis = 1)

    #右前脛骨筋
    A = R_Tib_Ant_A
    B = R_Tib_Ant_B
    C = R_Tib_Ant_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Tib_Ant = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Tib_Ant = pd.concat([df_R_Tib_Ant, df_condition, df_condition.describe()], axis = 1)

    #左内側腓腹筋
    A = L_Med_Gas_A
    B = L_Med_Gas_B
    C = L_Med_Gas_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Med_Gas = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Med_Gas = pd.concat([df_L_Med_Gas, df_condition, df_condition.describe()], axis = 1)

    #右内側腓腹筋
    A = R_Med_Gas_A
    B = R_Med_Gas_B
    C = R_Med_Gas_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_MEd_Gas = tukey_hsd(list("ABC"), (A, B, C))
    df_R_MEd_Gas = pd.concat([df_R_MEd_Gas, df_condition, df_condition.describe()], axis = 1)

    #Excel出力
    with pd.ExcelWriter("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx") as writer:
        df_L_Bic_Br.to_excel(writer, sheet_name = "左上腕二頭筋")
        df_R_Bic_Br.to_excel(writer, sheet_name = "右上腕二頭筋")
        df_L_Med_tri.to_excel(writer, sheet_name = "左上腕三頭筋")
        df_R_Med_tri.to_excel(writer, sheet_name = "右上腕三頭筋")
        df_L_Rec_Fem.to_excel(writer, sheet_name = "左大腿直筋")
        df_R_Rec_Fem.to_excel(writer, sheet_name = "右大腿直筋")
        df_L_Vlo.to_excel(writer, sheet_name = "左外側広筋")
        df_R_Vlo.to_excel(writer, sheet_name = "右外側広筋")
        df_L_Tib_Ant.to_excel(writer, sheet_name = "左前脛骨筋")
        df_R_Tib_Ant.to_excel(writer, sheet_name = "右前脛骨筋")
        df_L_Med_Gas.to_excel(writer, sheet_name = "左内側腓腹筋")
        df_R_MEd_Gas.to_excel(writer, sheet_name = "右内側腓腹筋")

    book = openpyxl.load_workbook("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx")
    sheet = book["左上腕二頭筋"]
    sheet["A1"] = "左上腕二頭筋"
    sheet = book["右上腕二頭筋"]
    sheet["A1"] = "右上腕二頭筋"
    sheet = book["左上腕三頭筋"]
    sheet["A1"] = "左上腕三頭筋"
    sheet = book["右上腕三頭筋"]
    sheet["A1"] = "右上腕三頭筋"
    sheet = book["左大腿直筋"]
    sheet["A1"] = "左大腿直筋"
    sheet = book["右大腿直筋"]
    sheet["A1"] = "右大腿直筋"
    sheet = book["左外側広筋"]
    sheet["A1"] = "左外側広筋"
    sheet = book["右外側広筋"]
    sheet["A1"] = "右外側広筋"
    sheet = book["左前脛骨筋"]
    sheet["A1"] = "左前脛骨筋"
    sheet = book["右前脛骨筋"]
    sheet["A1"] = "右前脛骨筋"
    sheet = book["左内側腓腹筋"]
    sheet["A1"] = "左内側腓腹筋"
    sheet = book["右内側腓腹筋"]
    sheet["A1"] = "右内側腓腹筋"

    book.save("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx")

#処理したい条件
conditionA = "condition5"
conditionB = "condition6"
conditionC = "condition7"

#データ個数
data_number = 4

for phase in tqdm(phase_list):

    print("・・・" + phase)

    phase_number = phase

    #Excelファイル名
    filename = phase_number + "_" + conditionA + conditionB + conditionC

    #条件ごとにファイル
    file_A = glob.glob(conditionA + "/" + "phase/" + phase_number + "/*.csv")
    file_B = glob.glob(conditionB + "/" + "phase/" + phase_number + "/*.csv")
    file_C = glob.glob(conditionC + "/" + "phase/" + phase_number + "/*.csv")

    #空リスト
    #conditionA
    L_Bic_Br_A = []
    R_Bic_Br_A = []
    L_Med_tri_A = []
    R_Med_tri_A = []
    L_Rec_Fem_A = []
    R_Rec_Fem_A = []
    L_Vlo_A = []
    R_Vlo_A = []
    L_Tib_Ant_A = []
    R_Tib_Ant_A = []
    L_Med_Gas_A = []
    R_Med_Gas_A = []
    #conditionB
    L_Bic_Br_B = []
    R_Bic_Br_B = []
    L_Med_tri_B = []
    R_Med_tri_B = []
    L_Rec_Fem_B = []
    R_Rec_Fem_B = []
    L_Vlo_B = []
    R_Vlo_B = []
    L_Tib_Ant_B = []
    R_Tib_Ant_B = []
    L_Med_Gas_B = []
    R_Med_Gas_B = []
    #conditionC
    L_Bic_Br_C = []
    R_Bic_Br_C = []
    L_Med_tri_C = []
    R_Med_tri_C = []
    L_Rec_Fem_C = []
    R_Rec_Fem_C = []
    L_Vlo_C = []
    R_Vlo_C = []
    L_Tib_Ant_C = []
    R_Tib_Ant_C = []
    L_Med_Gas_C = []
    R_Med_Gas_C = []

    print(conditionA)
    for file in tqdm(natsorted(file_A)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_A.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_A.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_A.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_A.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_A.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_A.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_A.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_A.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_A.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
            R_Tib_Ant_A.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_A.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_A.append(R_Med_Gas)

    #配列に変換
    L_Bic_Br_A = pd.Series(L_Bic_Br_A)
    R_Bic_Br_A = pd.Series(R_Bic_Br_A)
    L_Med_tri_A = pd.Series(L_Med_tri_A)
    R_Med_tri_A = pd.Series(R_Med_tri_A)
    L_Rec_Fem_A = pd.Series(L_Rec_Fem_A)
    R_Rec_Fem_A = pd.Series(R_Rec_Fem_A)
    L_Vlo_A = pd.Series(L_Vlo_A)
    R_Vlo_A = pd.Series(R_Vlo_A)
    L_Tib_Ant_A = pd.Series(L_Tib_Ant_A)
    R_Tib_Ant_A = pd.Series(R_Tib_Ant_A)
    L_Med_Gas_A = pd.Series(L_Med_Gas_A)
    R_Med_Gas_A = pd.Series(R_Med_Gas_A)

    print(conditionB)
    for file in tqdm(natsorted(file_B)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_B.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_B.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_B.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_B.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_B.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_B.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_B.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_B.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_B.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
            R_Tib_Ant_B.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_B.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_B.append(R_Med_Gas)

    #配列に変換
    L_Bic_Br_B = pd.Series(L_Bic_Br_B)
    R_Bic_Br_B = pd.Series(R_Bic_Br_B)
    L_Med_tri_B = pd.Series(L_Med_tri_B)
    R_Med_tri_B = pd.Series(R_Med_tri_B)
    L_Rec_Fem_B = pd.Series(L_Rec_Fem_B)
    R_Rec_Fem_B = pd.Series(R_Rec_Fem_B)
    L_Vlo_B = pd.Series(L_Vlo_B)
    R_Vlo_B = pd.Series(R_Vlo_B)
    L_Tib_Ant_B = pd.Series(L_Tib_Ant_B)
    R_Tib_Ant_B = pd.Series(R_Tib_Ant_B)
    L_Med_Gas_B = pd.Series(L_Med_Gas_B)
    R_Med_Gas_B = pd.Series(R_Med_Gas_B)

    print(conditionC)
    for file in tqdm(natsorted(file_C)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_C.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_C.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_C.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_C.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_C.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_C.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_C.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_C.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_C.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
            R_Tib_Ant_C.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_C.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_C.append(R_Med_Gas)

    #配列に変換
    L_Bic_Br_C = pd.Series(L_Bic_Br_C)
    R_Bic_Br_C = pd.Series(R_Bic_Br_C)
    L_Med_tri_C = pd.Series(L_Med_tri_C)
    R_Med_tri_C = pd.Series(R_Med_tri_C)
    L_Rec_Fem_C = pd.Series(L_Rec_Fem_C)
    R_Rec_Fem_C = pd.Series(R_Rec_Fem_C)
    L_Vlo_C = pd.Series(L_Vlo_C)
    R_Vlo_C = pd.Series(R_Vlo_C)
    L_Tib_Ant_C = pd.Series(L_Tib_Ant_C)
    R_Tib_Ant_C = pd.Series(R_Tib_Ant_C)
    L_Med_Gas_C = pd.Series(L_Med_Gas_C)
    R_Med_Gas_C = pd.Series(R_Med_Gas_C)

    #検定スタート
    #左上腕二頭筋
    A = L_Bic_Br_A
    B = L_Bic_Br_B
    C = L_Bic_Br_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Bic_Br = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Bic_Br = pd.concat([df_L_Bic_Br, df_condition, df_condition.describe()], axis = 1)

    #右上腕二頭筋
    A = R_Bic_Br_A
    B = R_Bic_Br_B
    C = R_Bic_Br_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Bic_Br = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Bic_Br = pd.concat([df_R_Bic_Br, df_condition, df_condition.describe()], axis = 1)

    #左上腕三頭筋
    A = L_Med_tri_A
    B = L_Med_tri_B
    C = L_Med_tri_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Med_tri = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Med_tri = pd.concat([df_L_Med_tri, df_condition, df_condition.describe()], axis = 1)

    #右上腕三頭筋
    A = R_Med_tri_A
    B = R_Med_tri_B
    C = R_Med_tri_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Med_tri = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Med_tri = pd.concat([df_R_Med_tri, df_condition, df_condition.describe()], axis = 1)

    #左大腿直筋
    A = L_Rec_Fem_A
    B = L_Rec_Fem_B
    C = L_Rec_Fem_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Rec_Fem = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Rec_Fem = pd.concat([df_L_Rec_Fem, df_condition, df_condition.describe()], axis = 1)

    #右大腿直筋
    A = R_Rec_Fem_A
    B = R_Rec_Fem_B
    C = R_Rec_Fem_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Rec_Fem = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Rec_Fem = pd.concat([df_R_Rec_Fem, df_condition, df_condition.describe()], axis = 1)

    #左外側広筋
    A = L_Vlo_A
    B = L_Vlo_B
    C = L_Vlo_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Vlo = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Vlo = pd.concat([df_L_Vlo, df_condition, df_condition.describe()], axis = 1)

    #右外側広筋
    A = R_Vlo_A
    B = R_Vlo_B
    C = R_Vlo_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Vlo = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Vlo = pd.concat([df_R_Vlo, df_condition, df_condition.describe()], axis = 1)

    #左前脛骨筋
    A = L_Tib_Ant_A
    B = L_Tib_Ant_B
    C = L_Tib_Ant_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Tib_Ant = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Tib_Ant = pd.concat([df_L_Tib_Ant, df_condition, df_condition.describe()], axis = 1)

    #右前脛骨筋
    A = R_Tib_Ant_A
    B = R_Tib_Ant_B
    C = R_Tib_Ant_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Tib_Ant = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Tib_Ant = pd.concat([df_R_Tib_Ant, df_condition, df_condition.describe()], axis = 1)

    #左内側腓腹筋
    A = L_Med_Gas_A
    B = L_Med_Gas_B
    C = L_Med_Gas_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Med_Gas = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Med_Gas = pd.concat([df_L_Med_Gas, df_condition, df_condition.describe()], axis = 1)

    #右内側腓腹筋
    A = R_Med_Gas_A
    B = R_Med_Gas_B
    C = R_Med_Gas_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_MEd_Gas = tukey_hsd(list("ABC"), (A, B, C))
    df_R_MEd_Gas = pd.concat([df_R_MEd_Gas, df_condition, df_condition.describe()], axis = 1)

    #Excel出力
    with pd.ExcelWriter("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx") as writer:
        df_L_Bic_Br.to_excel(writer, sheet_name = "左上腕二頭筋")
        df_R_Bic_Br.to_excel(writer, sheet_name = "右上腕二頭筋")
        df_L_Med_tri.to_excel(writer, sheet_name = "左上腕三頭筋")
        df_R_Med_tri.to_excel(writer, sheet_name = "右上腕三頭筋")
        df_L_Rec_Fem.to_excel(writer, sheet_name = "左大腿直筋")
        df_R_Rec_Fem.to_excel(writer, sheet_name = "右大腿直筋")
        df_L_Vlo.to_excel(writer, sheet_name = "左外側広筋")
        df_R_Vlo.to_excel(writer, sheet_name = "右外側広筋")
        df_L_Tib_Ant.to_excel(writer, sheet_name = "左前脛骨筋")
        df_R_Tib_Ant.to_excel(writer, sheet_name = "右前脛骨筋")
        df_L_Med_Gas.to_excel(writer, sheet_name = "左内側腓腹筋")
        df_R_MEd_Gas.to_excel(writer, sheet_name = "右内側腓腹筋")

    book = openpyxl.load_workbook("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx")
    sheet = book["左上腕二頭筋"]
    sheet["A1"] = "左上腕二頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右上腕二頭筋"]
    sheet["A1"] = "右上腕二頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左上腕三頭筋"]
    sheet["A1"] = "左上腕三頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右上腕三頭筋"]
    sheet["A1"] = "右上腕三頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左大腿直筋"]
    sheet["A1"] = "左大腿直筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右大腿直筋"]
    sheet["A1"] = "右大腿直筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左外側広筋"]
    sheet["A1"] = "左外側広筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右外側広筋"]
    sheet["A1"] = "右外側広筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左前脛骨筋"]
    sheet["A1"] = "左前脛骨筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右前脛骨筋"]
    sheet["A1"] = "右前脛骨筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左内側腓腹筋"]
    sheet["A1"] = "左内側腓腹筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右内側腓腹筋"]
    sheet["A1"] = "右内側腓腹筋"
    sheet.column_dimensions["A"].width = 13

    book.save("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx")    

#処理したい条件
conditionA = "condition8"
conditionB = "condition9"
conditionC = "condition10"

#データ個数
data_number = 4

for phase in tqdm(phase_list):
    
    print("・・・" + phase)

    phase_number = phase

    #Excelファイル名
    filename = phase_number + "_" + conditionA + conditionB + conditionC

    #条件ごとにファイル
    file_A = glob.glob(conditionA + "/" + "phase/" + phase_number + "/*.csv")
    file_B = glob.glob(conditionB + "/" + "phase/" + phase_number + "/*.csv")
    file_C = glob.glob(conditionC + "/" + "phase/" + phase_number + "/*.csv")

    #空リスト
    #conditionA
    L_Bic_Br_A = []
    R_Bic_Br_A = []
    L_Med_tri_A = []
    R_Med_tri_A = []
    L_Rec_Fem_A = []
    R_Rec_Fem_A = []
    L_Vlo_A = []
    R_Vlo_A = []
    L_Tib_Ant_A = []
    R_Tib_Ant_A = []
    L_Med_Gas_A = []
    R_Med_Gas_A = []
    #conditionB
    L_Bic_Br_B = []
    R_Bic_Br_B = []
    L_Med_tri_B = []
    R_Med_tri_B = []
    L_Rec_Fem_B = []
    R_Rec_Fem_B = []
    L_Vlo_B = []
    R_Vlo_B = []
    L_Tib_Ant_B = []
    R_Tib_Ant_B = []
    L_Med_Gas_B = []
    R_Med_Gas_B = []
    #conditionC
    L_Bic_Br_C = []
    R_Bic_Br_C = []
    L_Med_tri_C = []
    R_Med_tri_C = []
    L_Rec_Fem_C = []
    R_Rec_Fem_C = []
    L_Vlo_C = []
    R_Vlo_C = []
    L_Tib_Ant_C = []
    R_Tib_Ant_C = []
    L_Med_Gas_C = []
    R_Med_Gas_C = []

    print(conditionA)
    for file in tqdm(natsorted(file_A)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_A.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_A.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_A.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_A.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_A.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_A.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_A.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_A.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_A.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
             R_Tib_Ant_A.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_A.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_A.append(R_Med_Gas)

    #配列に変換
    L_Bic_Br_A = pd.Series(L_Bic_Br_A)
    R_Bic_Br_A = pd.Series(R_Bic_Br_A)
    L_Med_tri_A = pd.Series(L_Med_tri_A)
    R_Med_tri_A = pd.Series(R_Med_tri_A)
    L_Rec_Fem_A = pd.Series(L_Rec_Fem_A)
    R_Rec_Fem_A = pd.Series(R_Rec_Fem_A)
    L_Vlo_A = pd.Series(L_Vlo_A)
    R_Vlo_A = pd.Series(R_Vlo_A)
    L_Tib_Ant_A = pd.Series(L_Tib_Ant_A)
    R_Tib_Ant_A = pd.Series(R_Tib_Ant_A)
    L_Med_Gas_A = pd.Series(L_Med_Gas_A)
    R_Med_Gas_A = pd.Series(R_Med_Gas_A)

    print(conditionB)
    for file in tqdm(natsorted(file_B)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_B.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_B.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_B.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_B.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_B.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_B.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_B.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_B.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_B.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
             R_Tib_Ant_B.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_B.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_B.append(R_Med_Gas)
    #配列に変換
    L_Bic_Br_B = pd.Series(L_Bic_Br_B)
    R_Bic_Br_B = pd.Series(R_Bic_Br_B)
    L_Med_tri_B = pd.Series(L_Med_tri_B)
    R_Med_tri_B = pd.Series(R_Med_tri_B)
    L_Rec_Fem_B = pd.Series(L_Rec_Fem_B)
    R_Rec_Fem_B = pd.Series(R_Rec_Fem_B)
    L_Vlo_B = pd.Series(L_Vlo_B)
    R_Vlo_B = pd.Series(R_Vlo_B)
    L_Tib_Ant_B = pd.Series(L_Tib_Ant_B)
    R_Tib_Ant_B = pd.Series(R_Tib_Ant_B)
    L_Med_Gas_B = pd.Series(L_Med_Gas_B)
    R_Med_Gas_B = pd.Series(R_Med_Gas_B)

    print(conditionC)
    for file in tqdm(natsorted(file_C)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_C.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_C.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_C.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_C.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_C.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_C.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_C.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_C.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_C.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
             R_Tib_Ant_C.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_C.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_C.append(R_Med_Gas)

    #配列に変換
    L_Bic_Br_C = pd.Series(L_Bic_Br_C)
    R_Bic_Br_C = pd.Series(R_Bic_Br_C)
    L_Med_tri_C = pd.Series(L_Med_tri_C)
    R_Med_tri_C = pd.Series(R_Med_tri_C)
    L_Rec_Fem_C = pd.Series(L_Rec_Fem_C)
    R_Rec_Fem_C = pd.Series(R_Rec_Fem_C)
    L_Vlo_C = pd.Series(L_Vlo_C)
    R_Vlo_C = pd.Series(R_Vlo_C)
    L_Tib_Ant_C = pd.Series(L_Tib_Ant_C)
    R_Tib_Ant_C = pd.Series(R_Tib_Ant_C)
    L_Med_Gas_C = pd.Series(L_Med_Gas_C)
    R_Med_Gas_C = pd.Series(R_Med_Gas_C)

    #検定スタート
    #左上腕二頭筋
    A = L_Bic_Br_A
    B = L_Bic_Br_B
    C = L_Bic_Br_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Bic_Br = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Bic_Br = pd.concat([df_L_Bic_Br, df_condition, df_condition.describe()], axis = 1)

    #右上腕二頭筋
    A = R_Bic_Br_A
    B = R_Bic_Br_B
    C = R_Bic_Br_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Bic_Br = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Bic_Br = pd.concat([df_R_Bic_Br, df_condition, df_condition.describe()], axis = 1)

    #左上腕三頭筋
    A = L_Med_tri_A
    B = L_Med_tri_B
    C = L_Med_tri_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Med_tri = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Med_tri = pd.concat([df_L_Med_tri, df_condition, df_condition.describe()], axis = 1)

    #右上腕三頭筋
    A = R_Med_tri_A
    B = R_Med_tri_B
    C = R_Med_tri_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Med_tri = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Med_tri = pd.concat([df_R_Med_tri, df_condition, df_condition.describe()], axis = 1)

    #左大腿直筋
    A = L_Rec_Fem_A
    B = L_Rec_Fem_B
    C = L_Rec_Fem_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Rec_Fem = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Rec_Fem = pd.concat([df_L_Rec_Fem, df_condition, df_condition.describe()], axis = 1)

    #右大腿直筋
    A = R_Rec_Fem_A
    B = R_Rec_Fem_B
    C = R_Rec_Fem_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Rec_Fem = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Rec_Fem = pd.concat([df_R_Rec_Fem, df_condition, df_condition.describe()], axis = 1)

    #左外側広筋
    A = L_Vlo_A
    B = L_Vlo_B
    C = L_Vlo_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Vlo = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Vlo = pd.concat([df_L_Vlo, df_condition, df_condition.describe()], axis = 1)

    #右外側広筋
    A = R_Vlo_A
    B = R_Vlo_B
    C = R_Vlo_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Vlo = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Vlo = pd.concat([df_R_Vlo, df_condition, df_condition.describe()], axis = 1)

    #左前脛骨筋
    A = L_Tib_Ant_A
    B = L_Tib_Ant_B
    C = L_Tib_Ant_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Tib_Ant = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Tib_Ant = pd.concat([df_L_Tib_Ant, df_condition, df_condition.describe()], axis = 1)

    #右前脛骨筋
    A = R_Tib_Ant_A
    B = R_Tib_Ant_B
    C = R_Tib_Ant_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Tib_Ant = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Tib_Ant = pd.concat([df_R_Tib_Ant, df_condition, df_condition.describe()], axis = 1)

    #左内側腓腹筋
    A = L_Med_Gas_A
    B = L_Med_Gas_B
    C = L_Med_Gas_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Med_Gas = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Med_Gas = pd.concat([df_L_Med_Gas, df_condition, df_condition.describe()], axis = 1)

    #右内側腓腹筋
    A = R_Med_Gas_A
    B = R_Med_Gas_B
    C = R_Med_Gas_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_MEd_Gas = tukey_hsd(list("ABC"), (A, B, C))
    df_R_MEd_Gas = pd.concat([df_R_MEd_Gas, df_condition, df_condition.describe()], axis = 1)

    #Excel出力
    with pd.ExcelWriter("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx") as writer:
        df_L_Bic_Br.to_excel(writer, sheet_name = "左上腕二頭筋")
        df_R_Bic_Br.to_excel(writer, sheet_name = "右上腕二頭筋")
        df_L_Med_tri.to_excel(writer, sheet_name = "左上腕三頭筋")
        df_R_Med_tri.to_excel(writer, sheet_name = "右上腕三頭筋")
        df_L_Rec_Fem.to_excel(writer, sheet_name = "左大腿直筋")
        df_R_Rec_Fem.to_excel(writer, sheet_name = "右大腿直筋")
        df_L_Vlo.to_excel(writer, sheet_name = "左外側広筋")
        df_R_Vlo.to_excel(writer, sheet_name = "右外側広筋")
        df_L_Tib_Ant.to_excel(writer, sheet_name = "左前脛骨筋")
        df_R_Tib_Ant.to_excel(writer, sheet_name = "右前脛骨筋")
        df_L_Med_Gas.to_excel(writer, sheet_name = "左内側腓腹筋")
        df_R_MEd_Gas.to_excel(writer, sheet_name = "右内側腓腹筋")

    book = openpyxl.load_workbook("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx")
    sheet = book["左上腕二頭筋"]
    sheet["A1"] = "左上腕二頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右上腕二頭筋"]
    sheet["A1"] = "右上腕二頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左上腕三頭筋"]
    sheet["A1"] = "左上腕三頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右上腕三頭筋"]
    sheet["A1"] = "右上腕三頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左大腿直筋"]
    sheet["A1"] = "左大腿直筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右大腿直筋"]
    sheet["A1"] = "右大腿直筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左外側広筋"]
    sheet["A1"] = "左外側広筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右外側広筋"]
    sheet["A1"] = "右外側広筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左前脛骨筋"]
    sheet["A1"] = "左前脛骨筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右前脛骨筋"]
    sheet["A1"] = "右前脛骨筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左内側腓腹筋"]
    sheet["A1"] = "左内側腓腹筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右内側腓腹筋"]
    sheet["A1"] = "右内側腓腹筋"
    sheet.column_dimensions["A"].width = 13

    book.save("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx")    

#処理したい条件
conditionA = "condition2"
conditionB = "condition5"
conditionC = "condition8"

#データ個数
data_number = 4

for phase in tqdm(phase_list):
    
    print("・・・" + phase)

    phase_number = phase

    #Excelファイル名
    filename = phase_number + "_" + conditionA + conditionB + conditionC

    #条件ごとにファイル
    file_A = glob.glob(conditionA + "/" + "phase/" + phase_number + "/*.csv")
    file_B = glob.glob(conditionB + "/" + "phase/" + phase_number + "/*.csv")
    file_C = glob.glob(conditionC + "/" + "phase/" + phase_number + "/*.csv")

    #空リスト
    #conditionA
    L_Bic_Br_A = []
    R_Bic_Br_A = []
    L_Med_tri_A = []
    R_Med_tri_A = []
    L_Rec_Fem_A = []
    R_Rec_Fem_A = []
    L_Vlo_A = []
    R_Vlo_A = []
    L_Tib_Ant_A = []
    R_Tib_Ant_A = []
    L_Med_Gas_A = []
    R_Med_Gas_A = []
    #conditionB
    L_Bic_Br_B = []
    R_Bic_Br_B = []
    L_Med_tri_B = []
    R_Med_tri_B = []
    L_Rec_Fem_B = []
    R_Rec_Fem_B = []
    L_Vlo_B = []
    R_Vlo_B = []
    L_Tib_Ant_B = []
    R_Tib_Ant_B = []
    L_Med_Gas_B = []
    R_Med_Gas_B = []
    #conditionC
    L_Bic_Br_C = []
    R_Bic_Br_C = []
    L_Med_tri_C = []
    R_Med_tri_C = []
    L_Rec_Fem_C = []
    R_Rec_Fem_C = []
    L_Vlo_C = []
    R_Vlo_C = []
    L_Tib_Ant_C = []
    R_Tib_Ant_C = []
    L_Med_Gas_C = []
    R_Med_Gas_C = []

    print(conditionA)
    for file in tqdm(natsorted(file_A)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_A.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_A.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_A.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_A.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_A.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_A.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_A.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_A.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_A.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
             R_Tib_Ant_A.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_A.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_A.append(R_Med_Gas)
       
    #配列に変換
    L_Bic_Br_A = pd.Series(L_Bic_Br_A)
    R_Bic_Br_A = pd.Series(R_Bic_Br_A)
    L_Med_tri_A = pd.Series(L_Med_tri_A)
    R_Med_tri_A = pd.Series(R_Med_tri_A)
    L_Rec_Fem_A = pd.Series(L_Rec_Fem_A)
    R_Rec_Fem_A = pd.Series(R_Rec_Fem_A)
    L_Vlo_A = pd.Series(L_Vlo_A)
    R_Vlo_A = pd.Series(R_Vlo_A)
    L_Tib_Ant_A = pd.Series(L_Tib_Ant_A)
    R_Tib_Ant_A = pd.Series(R_Tib_Ant_A)
    L_Med_Gas_A = pd.Series(L_Med_Gas_A)
    R_Med_Gas_A = pd.Series(R_Med_Gas_A)

    print(conditionB)
    for file in tqdm(natsorted(file_B)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_B.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_B.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_B.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_B.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_B.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_B.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_B.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_B.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_B.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
             R_Tib_Ant_B.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_B.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_B.append(R_Med_Gas)

    #配列に変換
    L_Bic_Br_B = pd.Series(L_Bic_Br_B)
    R_Bic_Br_B = pd.Series(R_Bic_Br_B)
    L_Med_tri_B = pd.Series(L_Med_tri_B)
    R_Med_tri_B = pd.Series(R_Med_tri_B)
    L_Rec_Fem_B = pd.Series(L_Rec_Fem_B)
    R_Rec_Fem_B = pd.Series(R_Rec_Fem_B)
    L_Vlo_B = pd.Series(L_Vlo_B)
    R_Vlo_B = pd.Series(R_Vlo_B)
    L_Tib_Ant_B = pd.Series(L_Tib_Ant_B)
    R_Tib_Ant_B = pd.Series(R_Tib_Ant_B)
    L_Med_Gas_B = pd.Series(L_Med_Gas_B)
    R_Med_Gas_B = pd.Series(R_Med_Gas_B)

    print(conditionC)
    for file in tqdm(natsorted(file_C)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_C.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_C.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_C.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_C.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_C.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_C.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_C.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_C.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_C.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
             R_Tib_Ant_C.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_C.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_C.append(R_Med_Gas)

    #配列に変換
    L_Bic_Br_C = pd.Series(L_Bic_Br_C)
    R_Bic_Br_C = pd.Series(R_Bic_Br_C)
    L_Med_tri_C = pd.Series(L_Med_tri_C)
    R_Med_tri_C = pd.Series(R_Med_tri_C)
    L_Rec_Fem_C = pd.Series(L_Rec_Fem_C)
    R_Rec_Fem_C = pd.Series(R_Rec_Fem_C)
    L_Vlo_C = pd.Series(L_Vlo_C)
    R_Vlo_C = pd.Series(R_Vlo_C)
    L_Tib_Ant_C = pd.Series(L_Tib_Ant_C)
    R_Tib_Ant_C = pd.Series(R_Tib_Ant_C)
    L_Med_Gas_C = pd.Series(L_Med_Gas_C)
    R_Med_Gas_C = pd.Series(R_Med_Gas_C)

    #検定スタート
    #左上腕二頭筋
    A = L_Bic_Br_A
    B = L_Bic_Br_B
    C = L_Bic_Br_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Bic_Br = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Bic_Br = pd.concat([df_L_Bic_Br, df_condition, df_condition.describe()], axis = 1)

    #右上腕二頭筋
    A = R_Bic_Br_A
    B = R_Bic_Br_B
    C = R_Bic_Br_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Bic_Br = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Bic_Br = pd.concat([df_R_Bic_Br, df_condition, df_condition.describe()], axis = 1)

    #左上腕三頭筋
    A = L_Med_tri_A
    B = L_Med_tri_B
    C = L_Med_tri_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Med_tri = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Med_tri = pd.concat([df_L_Med_tri, df_condition, df_condition.describe()], axis = 1)

    #右上腕三頭筋
    A = R_Med_tri_A
    B = R_Med_tri_B
    C = R_Med_tri_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Med_tri = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Med_tri = pd.concat([df_R_Med_tri, df_condition, df_condition.describe()], axis = 1)

    #左大腿直筋
    A = L_Rec_Fem_A
    B = L_Rec_Fem_B
    C = L_Rec_Fem_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Rec_Fem = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Rec_Fem = pd.concat([df_L_Rec_Fem, df_condition, df_condition.describe()], axis = 1)

    #右大腿直筋
    A = R_Rec_Fem_A
    B = R_Rec_Fem_B
    C = R_Rec_Fem_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Rec_Fem = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Rec_Fem = pd.concat([df_R_Rec_Fem, df_condition, df_condition.describe()], axis = 1)

    #左外側広筋
    A = L_Vlo_A
    B = L_Vlo_B
    C = L_Vlo_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Vlo = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Vlo = pd.concat([df_L_Vlo, df_condition, df_condition.describe()], axis = 1)

    #右外側広筋
    A = R_Vlo_A
    B = R_Vlo_B
    C = R_Vlo_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Vlo = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Vlo = pd.concat([df_R_Vlo, df_condition, df_condition.describe()], axis = 1)

    #左前脛骨筋
    A = L_Tib_Ant_A
    B = L_Tib_Ant_B
    C = L_Tib_Ant_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Tib_Ant = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Tib_Ant = pd.concat([df_L_Tib_Ant, df_condition, df_condition.describe()], axis = 1)

    #右前脛骨筋
    A = R_Tib_Ant_A
    B = R_Tib_Ant_B
    C = R_Tib_Ant_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Tib_Ant = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Tib_Ant = pd.concat([df_R_Tib_Ant, df_condition, df_condition.describe()], axis = 1)

    #左内側腓腹筋
    A = L_Med_Gas_A
    B = L_Med_Gas_B
    C = L_Med_Gas_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Med_Gas = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Med_Gas = pd.concat([df_L_Med_Gas, df_condition, df_condition.describe()], axis = 1)

    #右内側腓腹筋
    A = R_Med_Gas_A
    B = R_Med_Gas_B
    C = R_Med_Gas_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_MEd_Gas = tukey_hsd(list("ABC"), (A, B, C))
    df_R_MEd_Gas = pd.concat([df_R_MEd_Gas, df_condition, df_condition.describe()], axis = 1)

    #Excel出力
    with pd.ExcelWriter("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx") as writer:
        df_L_Bic_Br.to_excel(writer, sheet_name = "左上腕二頭筋")
        df_R_Bic_Br.to_excel(writer, sheet_name = "右上腕二頭筋")
        df_L_Med_tri.to_excel(writer, sheet_name = "左上腕三頭筋")
        df_R_Med_tri.to_excel(writer, sheet_name = "右上腕三頭筋")
        df_L_Rec_Fem.to_excel(writer, sheet_name = "左大腿直筋")
        df_R_Rec_Fem.to_excel(writer, sheet_name = "右大腿直筋")
        df_L_Vlo.to_excel(writer, sheet_name = "左外側広筋")
        df_R_Vlo.to_excel(writer, sheet_name = "右外側広筋")
        df_L_Tib_Ant.to_excel(writer, sheet_name = "左前脛骨筋")
        df_R_Tib_Ant.to_excel(writer, sheet_name = "右前脛骨筋")
        df_L_Med_Gas.to_excel(writer, sheet_name = "左内側腓腹筋")
        df_R_MEd_Gas.to_excel(writer, sheet_name = "右内側腓腹筋")

    book = openpyxl.load_workbook("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx")
    sheet = book["左上腕二頭筋"]
    sheet["A1"] = "左上腕二頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右上腕二頭筋"]
    sheet["A1"] = "右上腕二頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左上腕三頭筋"]
    sheet["A1"] = "左上腕三頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右上腕三頭筋"]
    sheet["A1"] = "右上腕三頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左大腿直筋"]
    sheet["A1"] = "左大腿直筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右大腿直筋"]
    sheet["A1"] = "右大腿直筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左外側広筋"]
    sheet["A1"] = "左外側広筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右外側広筋"]
    sheet["A1"] = "右外側広筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左前脛骨筋"]
    sheet["A1"] = "左前脛骨筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右前脛骨筋"]
    sheet["A1"] = "右前脛骨筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左内側腓腹筋"]
    sheet["A1"] = "左内側腓腹筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右内側腓腹筋"]
    sheet["A1"] = "右内側腓腹筋"
    sheet.column_dimensions["A"].width = 13

    book.save("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx")    

#処理したい条件
conditionA = "condition3"
conditionB = "condition6"
conditionC = "condition9"

#データ個数
data_number = 4

for phase in tqdm(phase_list):
    
    print("・・・" + phase)

    phase_number = phase

    #Excelファイル名
    filename = phase_number + "_" + conditionA + conditionB + conditionC

    #条件ごとにファイル
    file_A = glob.glob(conditionA + "/" + "phase/" + phase_number + "/*.csv")
    file_B = glob.glob(conditionB + "/" + "phase/" + phase_number + "/*.csv")
    file_C = glob.glob(conditionC + "/" + "phase/" + phase_number + "/*.csv")

    #空リスト
    #conditionA
    L_Bic_Br_A = []
    R_Bic_Br_A = []
    L_Med_tri_A = []
    R_Med_tri_A = []
    L_Rec_Fem_A = []
    R_Rec_Fem_A = []
    L_Vlo_A = []
    R_Vlo_A = []
    L_Tib_Ant_A = []
    R_Tib_Ant_A = []
    L_Med_Gas_A = []
    R_Med_Gas_A = []
    #conditionB
    L_Bic_Br_B = []
    R_Bic_Br_B = []
    L_Med_tri_B = []
    R_Med_tri_B = []
    L_Rec_Fem_B = []
    R_Rec_Fem_B = []
    L_Vlo_B = []
    R_Vlo_B = []
    L_Tib_Ant_B = []
    R_Tib_Ant_B = []
    L_Med_Gas_B = []
    R_Med_Gas_B = []
    #conditionC
    L_Bic_Br_C = []
    R_Bic_Br_C = []
    L_Med_tri_C = []
    R_Med_tri_C = []
    L_Rec_Fem_C = []
    R_Rec_Fem_C = []
    L_Vlo_C = []
    R_Vlo_C = []
    L_Tib_Ant_C = []
    R_Tib_Ant_C = []
    L_Med_Gas_C = []
    R_Med_Gas_C = []

    print(conditionA)
    for file in tqdm(natsorted(file_A)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_A.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_A.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_A.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_A.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_A.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_A.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_A.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_A.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_A.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
             R_Tib_Ant_A.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_A.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_A.append(R_Med_Gas)
    #配列に変換
    L_Bic_Br_A = pd.Series(L_Bic_Br_A)
    R_Bic_Br_A = pd.Series(R_Bic_Br_A)
    L_Med_tri_A = pd.Series(L_Med_tri_A)
    R_Med_tri_A = pd.Series(R_Med_tri_A)
    L_Rec_Fem_A = pd.Series(L_Rec_Fem_A)
    R_Rec_Fem_A = pd.Series(R_Rec_Fem_A)
    L_Vlo_A = pd.Series(L_Vlo_A)
    R_Vlo_A = pd.Series(R_Vlo_A)
    L_Tib_Ant_A = pd.Series(L_Tib_Ant_A)
    R_Tib_Ant_A = pd.Series(R_Tib_Ant_A)
    L_Med_Gas_A = pd.Series(L_Med_Gas_A)
    R_Med_Gas_A = pd.Series(R_Med_Gas_A)

    print(conditionB)
    for file in tqdm(natsorted(file_B)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_B.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_B.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_B.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_B.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_B.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_B.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_B.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_B.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_B.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
             R_Tib_Ant_B.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_B.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_B.append(R_Med_Gas)

    #配列に変換
    L_Bic_Br_B = pd.Series(L_Bic_Br_B)
    R_Bic_Br_B = pd.Series(R_Bic_Br_B)
    L_Med_tri_B = pd.Series(L_Med_tri_B)
    R_Med_tri_B = pd.Series(R_Med_tri_B)
    L_Rec_Fem_B = pd.Series(L_Rec_Fem_B)
    R_Rec_Fem_B = pd.Series(R_Rec_Fem_B)
    L_Vlo_B = pd.Series(L_Vlo_B)
    R_Vlo_B = pd.Series(R_Vlo_B)
    L_Tib_Ant_B = pd.Series(L_Tib_Ant_B)
    R_Tib_Ant_B = pd.Series(R_Tib_Ant_B)
    L_Med_Gas_B = pd.Series(L_Med_Gas_B)
    R_Med_Gas_B = pd.Series(R_Med_Gas_B)

    print(conditionC)
    for file in tqdm(natsorted(file_C)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_C.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_C.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_C.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_C.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_C.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_C.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_C.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_C.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_C.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
             R_Tib_Ant_C.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_C.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_C.append(R_Med_Gas)

    #配列に変換
    L_Bic_Br_C = pd.Series(L_Bic_Br_C)
    R_Bic_Br_C = pd.Series(R_Bic_Br_C)
    L_Med_tri_C = pd.Series(L_Med_tri_C)
    R_Med_tri_C = pd.Series(R_Med_tri_C)
    L_Rec_Fem_C = pd.Series(L_Rec_Fem_C)
    R_Rec_Fem_C = pd.Series(R_Rec_Fem_C)
    L_Vlo_C = pd.Series(L_Vlo_C)
    R_Vlo_C = pd.Series(R_Vlo_C)
    L_Tib_Ant_C = pd.Series(L_Tib_Ant_C)
    R_Tib_Ant_C = pd.Series(R_Tib_Ant_C)
    L_Med_Gas_C = pd.Series(L_Med_Gas_C)
    R_Med_Gas_C = pd.Series(R_Med_Gas_C)

    #検定スタート
    #左上腕二頭筋
    A = L_Bic_Br_A
    B = L_Bic_Br_B
    C = L_Bic_Br_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Bic_Br = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Bic_Br = pd.concat([df_L_Bic_Br, df_condition, df_condition.describe()], axis = 1)

    #右上腕二頭筋
    A = R_Bic_Br_A
    B = R_Bic_Br_B
    C = R_Bic_Br_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Bic_Br = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Bic_Br = pd.concat([df_R_Bic_Br, df_condition, df_condition.describe()], axis = 1)

    #左上腕三頭筋
    A = L_Med_tri_A
    B = L_Med_tri_B
    C = L_Med_tri_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Med_tri = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Med_tri = pd.concat([df_L_Med_tri, df_condition, df_condition.describe()], axis = 1)

    #右上腕三頭筋
    A = R_Med_tri_A
    B = R_Med_tri_B
    C = R_Med_tri_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Med_tri = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Med_tri = pd.concat([df_R_Med_tri, df_condition, df_condition.describe()], axis = 1)

    #左大腿直筋
    A = L_Rec_Fem_A
    B = L_Rec_Fem_B
    C = L_Rec_Fem_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Rec_Fem = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Rec_Fem = pd.concat([df_L_Rec_Fem, df_condition, df_condition.describe()], axis = 1)

    #右大腿直筋
    A = R_Rec_Fem_A
    B = R_Rec_Fem_B
    C = R_Rec_Fem_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Rec_Fem = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Rec_Fem = pd.concat([df_R_Rec_Fem, df_condition, df_condition.describe()], axis = 1)

    #左外側広筋
    A = L_Vlo_A
    B = L_Vlo_B
    C = L_Vlo_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Vlo = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Vlo = pd.concat([df_L_Vlo, df_condition, df_condition.describe()], axis = 1)

    #右外側広筋
    A = R_Vlo_A
    B = R_Vlo_B
    C = R_Vlo_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Vlo = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Vlo = pd.concat([df_R_Vlo, df_condition, df_condition.describe()], axis = 1)

    #左前脛骨筋
    A = L_Tib_Ant_A
    B = L_Tib_Ant_B
    C = L_Tib_Ant_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Tib_Ant = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Tib_Ant = pd.concat([df_L_Tib_Ant, df_condition, df_condition.describe()], axis = 1)

    #右前脛骨筋
    A = R_Tib_Ant_A
    B = R_Tib_Ant_B
    C = R_Tib_Ant_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Tib_Ant = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Tib_Ant = pd.concat([df_R_Tib_Ant, df_condition, df_condition.describe()], axis = 1)

    #左内側腓腹筋
    A = L_Med_Gas_A
    B = L_Med_Gas_B
    C = L_Med_Gas_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Med_Gas = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Med_Gas = pd.concat([df_L_Med_Gas, df_condition, df_condition.describe()], axis = 1)

    #右内側腓腹筋
    A = R_Med_Gas_A
    B = R_Med_Gas_B
    C = R_Med_Gas_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_MEd_Gas = tukey_hsd(list("ABC"), (A, B, C))
    df_R_MEd_Gas = pd.concat([df_R_MEd_Gas, df_condition, df_condition.describe()], axis = 1)

    #Excel出力
    with pd.ExcelWriter("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx") as writer:
        df_L_Bic_Br.to_excel(writer, sheet_name = "左上腕二頭筋")
        df_R_Bic_Br.to_excel(writer, sheet_name = "右上腕二頭筋")
        df_L_Med_tri.to_excel(writer, sheet_name = "左上腕三頭筋")
        df_R_Med_tri.to_excel(writer, sheet_name = "右上腕三頭筋")
        df_L_Rec_Fem.to_excel(writer, sheet_name = "左大腿直筋")
        df_R_Rec_Fem.to_excel(writer, sheet_name = "右大腿直筋")
        df_L_Vlo.to_excel(writer, sheet_name = "左外側広筋")
        df_R_Vlo.to_excel(writer, sheet_name = "右外側広筋")
        df_L_Tib_Ant.to_excel(writer, sheet_name = "左前脛骨筋")
        df_R_Tib_Ant.to_excel(writer, sheet_name = "右前脛骨筋")
        df_L_Med_Gas.to_excel(writer, sheet_name = "左内側腓腹筋")
        df_R_MEd_Gas.to_excel(writer, sheet_name = "右内側腓腹筋")

    book = openpyxl.load_workbook("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx")
    sheet = book["左上腕二頭筋"]
    sheet["A1"] = "左上腕二頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右上腕二頭筋"]
    sheet["A1"] = "右上腕二頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左上腕三頭筋"]
    sheet["A1"] = "左上腕三頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右上腕三頭筋"]
    sheet["A1"] = "右上腕三頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左大腿直筋"]
    sheet["A1"] = "左大腿直筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右大腿直筋"]
    sheet["A1"] = "右大腿直筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左外側広筋"]
    sheet["A1"] = "左外側広筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右外側広筋"]
    sheet["A1"] = "右外側広筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左前脛骨筋"]
    sheet["A1"] = "左前脛骨筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右前脛骨筋"]
    sheet["A1"] = "右前脛骨筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左内側腓腹筋"]
    sheet["A1"] = "左内側腓腹筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右内側腓腹筋"]
    sheet["A1"] = "右内側腓腹筋"
    sheet.column_dimensions["A"].width = 13

    book.save("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx")    

#処理したい条件
conditionA = "condition4"
conditionB = "condition7"
conditionC = "condition10"

#データ個数
data_number = 4

for phase in tqdm(phase_list):
    
    print("・・・" + phase)

    phase_number = phase

    #Excelファイル名
    filename = phase_number + "_" + conditionA + conditionB + conditionC

    #条件ごとにファイル
    file_A = glob.glob(conditionA + "/" + "phase/" + phase_number + "/*.csv")
    file_B = glob.glob(conditionB + "/" + "phase/" + phase_number + "/*.csv")
    file_C = glob.glob(conditionC + "/" + "phase/" + phase_number + "/*.csv")

    #空リスト
    #conditionA
    L_Bic_Br_A = []
    R_Bic_Br_A = []
    L_Med_tri_A = []
    R_Med_tri_A = []
    L_Rec_Fem_A = []
    R_Rec_Fem_A = []
    L_Vlo_A = []
    R_Vlo_A = []
    L_Tib_Ant_A = []
    R_Tib_Ant_A = []
    L_Med_Gas_A = []
    R_Med_Gas_A = []
    #conditionB
    L_Bic_Br_B = []
    R_Bic_Br_B = []
    L_Med_tri_B = []
    R_Med_tri_B = []
    L_Rec_Fem_B = []
    R_Rec_Fem_B = []
    L_Vlo_B = []
    R_Vlo_B = []
    L_Tib_Ant_B = []
    R_Tib_Ant_B = []
    L_Med_Gas_B = []
    R_Med_Gas_B = []
    #conditionC
    L_Bic_Br_C = []
    R_Bic_Br_C = []
    L_Med_tri_C = []
    R_Med_tri_C = []
    L_Rec_Fem_C = []
    R_Rec_Fem_C = []
    L_Vlo_C = []
    R_Vlo_C = []
    L_Tib_Ant_C = []
    R_Tib_Ant_C = []
    L_Med_Gas_C = []
    R_Med_Gas_C = []

    print(conditionA)
    for file in tqdm(natsorted(file_A)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_A.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_A.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_A.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_A.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_A.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_A.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_A.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_A.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_A.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
             R_Tib_Ant_A.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_A.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_A.append(R_Med_Gas)

    #配列に変換
    L_Bic_Br_A = pd.Series(L_Bic_Br_A)
    R_Bic_Br_A = pd.Series(R_Bic_Br_A)
    L_Med_tri_A = pd.Series(L_Med_tri_A)
    R_Med_tri_A = pd.Series(R_Med_tri_A)
    L_Rec_Fem_A = pd.Series(L_Rec_Fem_A)
    R_Rec_Fem_A = pd.Series(R_Rec_Fem_A)
    L_Vlo_A = pd.Series(L_Vlo_A)
    R_Vlo_A = pd.Series(R_Vlo_A)
    L_Tib_Ant_A = pd.Series(L_Tib_Ant_A)
    R_Tib_Ant_A = pd.Series(R_Tib_Ant_A)
    L_Med_Gas_A = pd.Series(L_Med_Gas_A)
    R_Med_Gas_A = pd.Series(R_Med_Gas_A)

    print(conditionB)
    for file in tqdm(natsorted(file_B)):
        df = pd.read_csv(file)
        #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_B.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_B.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_B.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_B.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_B.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_B.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_B.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_B.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_B.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
             R_Tib_Ant_B.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_B.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_B.append(R_Med_Gas)

    #配列に変換
    L_Bic_Br_B = pd.Series(L_Bic_Br_B)
    R_Bic_Br_B = pd.Series(R_Bic_Br_B)
    L_Med_tri_B = pd.Series(L_Med_tri_B)
    R_Med_tri_B = pd.Series(R_Med_tri_B)
    L_Rec_Fem_B = pd.Series(L_Rec_Fem_B)
    R_Rec_Fem_B = pd.Series(R_Rec_Fem_B)
    L_Vlo_B = pd.Series(L_Vlo_B)
    R_Vlo_B = pd.Series(R_Vlo_B)
    L_Tib_Ant_B = pd.Series(L_Tib_Ant_B)
    R_Tib_Ant_B = pd.Series(R_Tib_Ant_B)
    L_Med_Gas_B = pd.Series(L_Med_Gas_B)
    R_Med_Gas_B = pd.Series(R_Med_Gas_B)

    print(conditionC)
    for file in tqdm(natsorted(file_C)):
        df = pd.read_csv(file)
       #平均算出
        try:
            L_Bic_Br = df["L.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            L_Bic_Br_C.append(L_Bic_Br)
        try:
            R_Bic_Br = df["R.Biceps Br.(uV):"].mean()
        except:
            pass
        else:
            R_Bic_Br_C.append(R_Bic_Br)
        try:    
            L_Med_tri = df["L.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            L_Med_tri_C.append(L_Med_tri)
        try:    
            R_Med_tri = df["R.Med. Triceps(uV):"].mean()
        except:
            pass
        else:
            R_Med_tri_C.append(R_Med_tri)
        try:
            L_Rec_Fem = df["L.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            L_Rec_Fem_C.append(L_Rec_Fem)
        try:    
            R_Rec_Fem  =df["R.Rectus Fem.(uV):"].mean()
        except:
            pass
        else:
            R_Rec_Fem_C.append(R_Rec_Fem)
        try:    
            L_Vlo = df["L.Vlo(uV):"].mean()
        except:
            pass
        else:
            L_Vlo_C.append(L_Vlo)
        try:
            R_Vlo = df["R.Vlo(uV):"].mean()
        except:
            pass    
        else:
            R_Vlo_C.append(R_Vlo)
        try:
            L_Tib_Ant = df["L.Tib.Ant.(uV):"].mean()
        except:
            pass    
        else:
            L_Tib_Ant_C.append(L_Tib_Ant)
        try:    
            R_Tib_Ant = df["R.Tib.Ant.(uV):"].mean()
        except:
            pass
        else:
             R_Tib_Ant_C.append(R_Tib_Ant)
        try:
            L_Med_Gas = df["L.Med. Gastro(uV):"].mean()
        except:
            pass
        else:
            L_Med_Gas_C.append(L_Med_Gas)
        try:    
            R_Med_Gas = df["R.Med. Gastro(uV):"].mean()
        except:
            pass    
        else:
            R_Med_Gas_C.append(R_Med_Gas)

    #配列に変換
    L_Bic_Br_C = pd.Series(L_Bic_Br_C)
    R_Bic_Br_C = pd.Series(R_Bic_Br_C)
    L_Med_tri_C = pd.Series(L_Med_tri_C)
    R_Med_tri_C = pd.Series(R_Med_tri_C)
    L_Rec_Fem_C = pd.Series(L_Rec_Fem_C)
    R_Rec_Fem_C = pd.Series(R_Rec_Fem_C)
    L_Vlo_C = pd.Series(L_Vlo_C)
    R_Vlo_C = pd.Series(R_Vlo_C)
    L_Tib_Ant_C = pd.Series(L_Tib_Ant_C)
    R_Tib_Ant_C = pd.Series(R_Tib_Ant_C)
    L_Med_Gas_C = pd.Series(L_Med_Gas_C)
    R_Med_Gas_C = pd.Series(R_Med_Gas_C)

    #検定スタート
    #左上腕二頭筋
    A = L_Bic_Br_A
    B = L_Bic_Br_B
    C = L_Bic_Br_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Bic_Br = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Bic_Br = pd.concat([df_L_Bic_Br, df_condition, df_condition.describe()], axis = 1)

    #右上腕二頭筋
    A = R_Bic_Br_A
    B = R_Bic_Br_B
    C = R_Bic_Br_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Bic_Br = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Bic_Br = pd.concat([df_R_Bic_Br, df_condition, df_condition.describe()], axis = 1)

    #左上腕三頭筋
    A = L_Med_tri_A
    B = L_Med_tri_B
    C = L_Med_tri_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Med_tri = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Med_tri = pd.concat([df_L_Med_tri, df_condition, df_condition.describe()], axis = 1)

    #右上腕三頭筋
    A = R_Med_tri_A
    B = R_Med_tri_B
    C = R_Med_tri_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Med_tri = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Med_tri = pd.concat([df_R_Med_tri, df_condition, df_condition.describe()], axis = 1)

    #左大腿直筋
    A = L_Rec_Fem_A
    B = L_Rec_Fem_B
    C = L_Rec_Fem_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Rec_Fem = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Rec_Fem = pd.concat([df_L_Rec_Fem, df_condition, df_condition.describe()], axis = 1)

    #右大腿直筋
    A = R_Rec_Fem_A
    B = R_Rec_Fem_B
    C = R_Rec_Fem_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Rec_Fem = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Rec_Fem = pd.concat([df_R_Rec_Fem, df_condition, df_condition.describe()], axis = 1)

    #左外側広筋
    A = L_Vlo_A
    B = L_Vlo_B
    C = L_Vlo_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Vlo = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Vlo = pd.concat([df_L_Vlo, df_condition, df_condition.describe()], axis = 1)

    #右外側広筋
    A = R_Vlo_A
    B = R_Vlo_B
    C = R_Vlo_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Vlo = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Vlo = pd.concat([df_R_Vlo, df_condition, df_condition.describe()], axis = 1)

    #左前脛骨筋
    A = L_Tib_Ant_A
    B = L_Tib_Ant_B
    C = L_Tib_Ant_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Tib_Ant = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Tib_Ant = pd.concat([df_L_Tib_Ant, df_condition, df_condition.describe()], axis = 1)

    #右前脛骨筋
    A = R_Tib_Ant_A
    B = R_Tib_Ant_B
    C = R_Tib_Ant_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_Tib_Ant = tukey_hsd(list("ABC"), (A, B, C))
    df_R_Tib_Ant = pd.concat([df_R_Tib_Ant, df_condition, df_condition.describe()], axis = 1)

    #左内側腓腹筋
    A = L_Med_Gas_A
    B = L_Med_Gas_B
    C = L_Med_Gas_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_L_Med_Gas = tukey_hsd(list("ABC"), (A, B, C))
    df_L_Med_Gas = pd.concat([df_L_Med_Gas, df_condition, df_condition.describe()], axis = 1)

    #右内側腓腹筋
    A = R_Med_Gas_A
    B = R_Med_Gas_B
    C = R_Med_Gas_C

    df_condition = pd.DataFrame({conditionA:A, conditionB:B, conditionC:C})
    df_R_MEd_Gas = tukey_hsd(list("ABC"), (A, B, C))
    df_R_MEd_Gas = pd.concat([df_R_MEd_Gas, df_condition, df_condition.describe()], axis = 1)

    #Excel出力
    with pd.ExcelWriter("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx") as writer:
        df_L_Bic_Br.to_excel(writer, sheet_name = "左上腕二頭筋")
        df_R_Bic_Br.to_excel(writer, sheet_name = "右上腕二頭筋")
        df_L_Med_tri.to_excel(writer, sheet_name = "左上腕三頭筋")
        df_R_Med_tri.to_excel(writer, sheet_name = "右上腕三頭筋")
        df_L_Rec_Fem.to_excel(writer, sheet_name = "左大腿直筋")
        df_R_Rec_Fem.to_excel(writer, sheet_name = "右大腿直筋")
        df_L_Vlo.to_excel(writer, sheet_name = "左外側広筋")
        df_R_Vlo.to_excel(writer, sheet_name = "右外側広筋")
        df_L_Tib_Ant.to_excel(writer, sheet_name = "左前脛骨筋")
        df_R_Tib_Ant.to_excel(writer, sheet_name = "右前脛骨筋")
        df_L_Med_Gas.to_excel(writer, sheet_name = "左内側腓腹筋")
        df_R_MEd_Gas.to_excel(writer, sheet_name = "右内側腓腹筋")

    book = openpyxl.load_workbook("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx")
    sheet = book["左上腕二頭筋"]
    sheet["A1"] = "左上腕二頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右上腕二頭筋"]
    sheet["A1"] = "右上腕二頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左上腕三頭筋"]
    sheet["A1"] = "左上腕三頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右上腕三頭筋"]
    sheet["A1"] = "右上腕三頭筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左大腿直筋"]
    sheet["A1"] = "左大腿直筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右大腿直筋"]
    sheet["A1"] = "右大腿直筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左外側広筋"]
    sheet["A1"] = "左外側広筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右外側広筋"]
    sheet["A1"] = "右外側広筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左前脛骨筋"]
    sheet["A1"] = "左前脛骨筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右前脛骨筋"]
    sheet["A1"] = "右前脛骨筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["左内側腓腹筋"]
    sheet["A1"] = "左内側腓腹筋"
    sheet.column_dimensions["A"].width = 13
    sheet = book["右内側腓腹筋"]
    sheet["A1"] = "右内側腓腹筋"
    sheet.column_dimensions["A"].width = 13

    book.save("Excel_EMG_Tukey/" + phase_number + "/" + filename + ".xlsx")

print("終了")